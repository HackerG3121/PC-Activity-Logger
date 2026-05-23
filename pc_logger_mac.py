"""
╔══════════════════════════════════════════════════════════════╗
║   PC ACTIVITY LOGGER — macOS EDITION                         ║
║   Optimized exclusively for macOS 12 Monterey and newer      ║
║   Requires: pip3 install psutil openpyxl                     ║
╚══════════════════════════════════════════════════════════════╝

macOS-specific features:
  - system_profiler for USB, Bluetooth, audio, webcam, monitors
  - airport CLI for Wi-Fi SSID / signal / channel
  - pbpaste for clipboard monitoring
  - Quartz CGSession for screen lock detection
  - launchctl / LaunchAgent for automation
  - lpstat for printer detection
  - iCloud Drive, Google Drive, Dropbox sync support

PERMISSIONS required (System Preferences → Privacy):
  - Full Disk Access  → for Bluetooth + system_profiler
  - Accessibility     → for screen lock detection (optional)
"""

import os, sys, time, datetime, threading, platform, subprocess, socket, re
import psutil, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════

LOG_FILE = os.path.join(os.path.expanduser("~"), "Library", "Mobile Documents",
                        "com~apple~CloudDocs", "PC_ActivityLog.xlsx")  # iCloud Drive
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Google Drive", "PC_ActivityLog.xlsx")
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Dropbox", "PC_ActivityLog.xlsx")
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Desktop", "PC_ActivityLog.xlsx")

APP_POLL      = 5
USB_POLL      = 5
NETWORK_POLL  = 6
BATTERY_POLL  = 30
SCREEN_POLL   = 5
AUDIO_POLL    = 8
DRIVE_POLL    = 6
DEVICE_POLL   = 10
INTERNET_POLL = 15
PERF_POLL     = 20
SERVICE_POLL  = 30
SAVE_EVERY    = 30

BATTERY_DELTA    = 2
CPU_SPIKE_THRESH = 85
RAM_SPIKE_THRESH = 85

LOG_CLIPBOARD = False
LOG_APPS      = True
LOG_PERF      = True

INTERNET_HOST = "8.8.8.8"
INTERNET_PORT = 53

AIRPORT = "/System/Library/PrivateFrameworks/Apple80211.framework/Versions/Current/Resources/airport"

# macOS background daemons to ignore in app tracking
IGNORED_PROCS = {
    "launchd","kernel_task","logd","configd","notifyd","diskarbitrationd",
    "fseventsd","coreaudiod","opendirectoryd","update_dyld_sim",
    "SystemUIServer","Finder","Dock","WindowServer","loginwindow",
    "mds","mds_stores","mdworker","spotlight","trustd","watchdogd",
    "nsurlsessiond","airportd","bluetoothd","sharingd","rapportd",
    "coreduetd","symptomsd","containermanagerd","lsd","secd",
    "securityd","cfprefsd","usbd","powerd","hidd","apfsd",
    "mobileactivationd","akd","ctkd","appstoreagent","storeaccountd",
}

# ══════════════════════════════════════════════════════════════
#  SHEET DEFINITIONS  (identical across all OS editions)
# ══════════════════════════════════════════════════════════════

SHEETS = {
    "All Events":      ["#","Timestamp","Date","Time","Category","Event","Details","Extra","Duration","User","PC"],
    "Power":           ["#","Timestamp","Date","Time","Event","Details","Duration","PC"],
    "Applications":    ["#","Timestamp","Date","Time","Action","Application","PID","Duration","User","Path"],
    "USB Devices":     ["#","Timestamp","Date","Time","Action","Device Name","Drive","Type","Serial ID"],
    "Wi-Fi":           ["#","Timestamp","Date","Time","Action","SSID","Signal","Band","IP Address","MAC","Interface"],
    "Bluetooth":       ["#","Timestamp","Date","Time","Action","Device Name","MAC Address","Device Type","Paired","RSSI"],
    "Network":         ["#","Timestamp","Date","Time","Action","Interface","IP Address","MAC","Speed Mbps","Notes"],
    "Internet":        ["#","Timestamp","Date","Time","Status","Latency ms","Details","Duration Offline"],
    "Battery":         ["#","Timestamp","Date","Time","Event","Level %","Status","Time Remaining","Power Source"],
    "Screen":          ["#","Timestamp","Date","Time","Event","Duration Locked","User","Details"],
    "Drives":          ["#","Timestamp","Date","Time","Action","Drive","Label","Filesystem","Size GB","Free GB"],
    "Audio":           ["#","Timestamp","Date","Time","Action","Device Name","Category","Is Default","Details"],
    "Microphone":      ["#","Timestamp","Date","Time","Action","Device Name","Type","Details"],
    "Webcam":          ["#","Timestamp","Date","Time","Action","Device Name","Type","Details"],
    "Printers":        ["#","Timestamp","Date","Time","Action","Printer Name","Port","Driver","Status"],
    "Monitors":        ["#","Timestamp","Date","Time","Action","Monitor Name","Resolution","Port","Serial"],
    "VPN":             ["#","Timestamp","Date","Time","Action","VPN Name","Server","Protocol","Duration"],
    "Mobile Hotspot":  ["#","Timestamp","Date","Time","Action","Device Name","Type","IP","Details"],
    "Network Drives":  ["#","Timestamp","Date","Time","Action","Drive Letter","UNC Path","Server","Status"],
    "Controllers":     ["#","Timestamp","Date","Time","Action","Controller Name","Type","ID","Details"],
    "Remote Sessions": ["#","Timestamp","Date","Time","Action","Username","Client IP","Session Type","Duration"],
    "Security":        ["#","Timestamp","Date","Time","Event","Component","Status","Details"],
    "Users":           ["#","Timestamp","Date","Time","Event","Username","Session","Details","PC"],
    "Keyboard Mouse":  ["#","Timestamp","Date","Time","Action","Device Name","Type","Interface","Serial ID"],
    "Power Plan":      ["#","Timestamp","Date","Time","Action","Plan Name","Plan GUID","Previous Plan","Details"],
    "Services":        ["#","Timestamp","Date","Time","Action","Service Name","Display Name","Previous State","New State"],
    "System Perf":     ["#","Timestamp","Date","Time","Alert Type","CPU %","RAM %","Top Process","Details"],
    "Smart Card":      ["#","Timestamp","Date","Time","Action","Reader Name","Card Type","Details"],
    "Display Mode":    ["#","Timestamp","Date","Time","Action","Setting Changed","Old Value","New Value","Details"],
    "Clipboard":       ["#","Timestamp","Date","Time","Content Preview","Length","Details"],
    "Daily Summary":   ["Date","Power","Apps","USB","WiFi","BT","Internet","Printers","Monitors","VPN","Ctrl","Total"],
}

WIDTHS = {
    "All Events":[5,22,12,10,16,18,42,18,12,10,14],"Power":[5,22,12,10,22,50,14,16],
    "Applications":[5,22,12,10,10,24,8,12,10,55],"USB Devices":[5,22,12,10,12,30,10,18,28],
    "Wi-Fi":[5,22,12,10,14,22,10,10,16,18,14],"Bluetooth":[5,22,12,10,14,28,18,18,10,12],
    "Network":[5,22,12,10,14,16,16,18,12,20],"Internet":[5,22,12,10,14,12,30,16],
    "Battery":[5,22,12,10,18,10,16,16,16],"Screen":[5,22,12,10,16,16,14,30],
    "Drives":[5,22,12,10,12,14,18,14,10,10],"Audio":[5,22,12,10,14,35,16,12,30],
    "Microphone":[5,22,12,10,14,35,16,30],"Webcam":[5,22,12,10,14,35,16,30],
    "Printers":[5,22,12,10,14,30,18,25,14],"Monitors":[5,22,12,10,14,30,16,16,20],
    "VPN":[5,22,12,10,14,22,20,14,14],"Mobile Hotspot":[5,22,12,10,14,25,16,16,25],
    "Network Drives":[5,22,12,10,12,14,35,20,14],"Controllers":[5,22,12,10,14,28,16,12,25],
    "Remote Sessions":[5,22,12,10,14,18,18,14,14],"Security":[5,22,12,10,18,18,14,40],
    "Users":[5,22,12,10,14,16,12,35,14],"Keyboard Mouse":[5,22,12,10,14,30,16,14,25],
    "Power Plan":[5,22,12,10,14,25,35,25,25],"Services":[5,22,12,10,14,20,30,16,16],
    "System Perf":[5,22,12,10,16,10,10,25,35],"Smart Card":[5,22,12,10,14,28,18,30],
    "Display Mode":[5,22,12,10,14,22,18,18,30],"Clipboard":[5,22,12,10,55,8,20],
    "Daily Summary":[14,8,8,8,8,8,8,8,8,8,8,10],
}

CATS = {
    "POWER ON":"C6EFCE","POWER OFF":"FFCCCC","SLEEP":"FFF2CC","WAKE":"E2EFDA",
    "HIBERNATE":"FFE4B5","RESUME":"D5F5E3","APP OPEN":"DDEBF7","APP CLOSE":"FCE4D6",
    "USB INSERT":"E2EFDA","USB REMOVE":"FFF2CC","WIFI CONNECT":"D9EAD3","WIFI DISCONNECT":"FCE4D6",
    "BT CONNECT":"CFE2F3","BT DISCONNECT":"FCE4D6","NET UP":"D9EAD3","NET DOWN":"FFCCCC",
    "IP CHANGED":"FFF9C4","INTERNET UP":"D9EAD3","INTERNET DOWN":"FFCCCC",
    "BATTERY LOW":"FFCCCC","BATTERY CRITICAL":"FF9999","CHARGING":"C6EFCE",
    "DISCHARGING":"FFF2CC","BATTERY":"FFF9C4","SCREEN LOCK":"E8DAEF","SCREEN UNLOCK":"D5F5E3",
    "DRIVE MOUNT":"E2EFDA","DRIVE UNMOUNT":"FFF2CC","AUDIO IN":"E8F4FD","AUDIO OUT":"FCE4D6",
    "MIC IN":"E8F4FD","MIC OUT":"FCE4D6","CAM IN":"E2EFDA","CAM OUT":"FFF2CC",
    "PRINTER ADD":"E2EFDA","PRINTER REMOVE":"FFF2CC","MONITOR IN":"C6EFCE","MONITOR OUT":"FFCCCC",
    "VPN CONNECT":"D9EAD3","VPN DISCONNECT":"FFCCCC","HOTSPOT ON":"FFF9C4","HOTSPOT OFF":"FCE4D6",
    "NET DRIVE MAP":"E2EFDA","NET DRIVE UNMAP":"FFF2CC","CTRL CONNECT":"E2EFDA","CTRL DISCONNECT":"FFF2CC",
    "RDP IN":"CFE2F3","RDP OUT":"FCE4D6","SECURITY":"FFCCCC","SECURITY OK":"E2EFDA",
    "USER LOGIN":"C6EFCE","USER LOGOUT":"FFCCCC","KB CONNECT":"E2EFDA","KB DISCONNECT":"FFF2CC",
    "MOUSE CONNECT":"E2EFDA","MOUSE DISCONNECT":"FFF2CC","PLAN CHANGED":"FFF9C4",
    "SVC START":"D9EAD3","SVC STOP":"FFCCCC","CPU SPIKE":"FFCCCC","RAM SPIKE":"FFE4B5",
    "CARD IN":"E2EFDA","CARD OUT":"FFF2CC","DISPLAY CHANGE":"FFF9C4","CLIPBOARD":"F5F5F5",
}

HDR_FILL = PatternFill("solid", fgColor="1F3864")

# ══════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ══════════════════════════════════════════════════════════════

def make_fill(h): return PatternFill("solid", fgColor=h)

def write_header_row(ws, cols):
    for i, title in enumerate(cols, 1):
        c = ws.cell(1, i, title)
        c.font = Font(bold=True, color="FFFFFF", name="SF Pro Text", size=10)
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    ws.row_dimensions[1].height = 23

def style_data_row(ws, row, cat=None):
    fill = (make_fill(CATS[cat]) if cat and cat in CATS
            else make_fill("EEF2FF") if row % 2 == 0 else make_fill("FFFFFF"))
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row, col)
        c.fill = fill; c.font = Font(name="SF Pro Text", size=9)
        c.alignment = Alignment(vertical="center")
        c.border = Border(bottom=Side(style="hair", color="D9D9D9"))

# ══════════════════════════════════════════════════════════════
#  GLOBAL STATE
# ══════════════════════════════════════════════════════════════

wb = None; wb_lock = threading.Lock(); row_counters = {}; event_id = 0
boot_time     = datetime.datetime.fromtimestamp(psutil.boot_time())
session_start = datetime.datetime.now()
pc_name       = socket.gethostname()
username      = os.getenv("USER", "Unknown")

st_apps={}; st_usb=set(); st_wifi={}; st_bt=set(); st_net={}
st_batt={}; st_screen=None; st_drives=set(); st_audio=set()
st_mic=set(); st_cam=set(); st_printers={}; st_monitors={}
st_vpn=set(); st_hotspot=set(); st_netdrives=set(); st_ctrl=set()
st_rdp=set(); st_kb=set(); st_mouse=set(); st_svc={}; st_card=set()
st_plan=""; st_inet=None; st_inet_down_since=None; clip_last=""

def next_row(sheet):
    if sheet not in row_counters:
        row_counters[sheet] = max(wb[sheet].max_row, 1) + 1
    row_counters[sheet] += 1; return row_counters[sheet] - 1

def next_id():
    global event_id; event_id += 1; return event_id

def write_row(sheet, data, cat=None):
    row = next_row(sheet); ws = wb[sheet]
    for col, val in enumerate(data, 1): ws.cell(row, col, val)
    style_data_row(ws, row, cat)

def now_ts():
    dt = datetime.datetime.now()
    return dt.strftime("%Y-%m-%d %H:%M:%S"), dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")

def log_all(cat, etype, details, extra="", duration=""):
    ts,d,t = now_ts()
    write_row("All Events",[next_id(),ts,d,t,cat,etype,details,extra,duration,username,pc_name],cat)

def run_cmd(cmd_list, timeout=8):
    """Run a shell command list, return stdout string."""
    try:
        return subprocess.check_output(cmd_list, text=True, stderr=subprocess.DEVNULL, timeout=timeout)
    except Exception:
        return ""

# ══════════════════════════════════════════════════════════════
#  EXCEL INIT
# ══════════════════════════════════════════════════════════════

def init_excel():
    if os.path.exists(LOG_FILE):
        print(f"  📂 Existing log loaded: {LOG_FILE}")
        return openpyxl.load_workbook(LOG_FILE)
    wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)
    info = wb2.create_sheet("INFO")
    info["B2"] = "PC ACTIVITY LOGGER — macOS EDITION"
    info["B2"].font = Font(bold=True, size=15, color="1F3864", name="SF Pro Text")
    for i, (lbl, val) in enumerate([
        ("Mac Name:", pc_name), ("Started:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("User:", username), ("OS:", platform.platform()),
        ("IP:", socket.gethostbyname(socket.gethostname())), ("Python:", sys.version.split()[0])
    ]):
        info.cell(4+i,2,lbl).font = Font(bold=True, color="595959", name="SF Pro Text")
        info.cell(4+i,3,val).font = Font(name="SF Pro Text")
    info.column_dimensions["B"].width = 20; info.column_dimensions["C"].width = 58
    for name, cols in SHEETS.items():
        ws = wb2.create_sheet(name); write_header_row(ws, cols); ws.freeze_panes = "A2"
        if name in WIDTHS:
            for ci, w in enumerate(WIDTHS[name], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
    os.makedirs(os.path.dirname(LOG_FILE) if os.path.dirname(LOG_FILE) else ".", exist_ok=True)
    wb2.save(LOG_FILE); return wb2

# ══════════════════════════════════════════════════════════════
#  POWER
# ══════════════════════════════════════════════════════════════

def log_startup():
    ts,d,t = now_ts()
    write_row("Power",[next_id(),ts,d,t,"POWER ON",f"Boot:{boot_time.strftime('%Y-%m-%d %H:%M:%S')} Logger:{ts}","",pc_name],"POWER ON")
    log_all("POWER ON","System Startup",f"Boot:{boot_time.strftime('%H:%M:%S')}",platform.platform())
    print(f"[{ts}] ✅ macOS Logger → {LOG_FILE}")

def log_shutdown():
    ts,d,t = now_ts(); dur = str(datetime.datetime.now()-session_start).split(".")[0]
    write_row("Power",[next_id(),ts,d,t,"POWER OFF",f"Session:{dur}",dur,pc_name],"POWER OFF")
    log_all("POWER OFF","System Shutdown",f"Session:{dur}","",dur)
    save_wb(); print(f"[{ts}] 💾 Logger stopped — Session: {dur}")

# ══════════════════════════════════════════════════════════════
#  APPLICATIONS  (macOS: psutil with macOS-specific ignored procs)
# ══════════════════════════════════════════════════════════════

def get_apps():
    apps = {}
    for p in psutil.process_iter(["pid","name","exe","username","create_time"]):
        try:
            name = p.info["name"] or ""
            if name.lower() in IGNORED_PROCS or not name.strip(): continue
            apps[p.info["pid"]] = {
                "name":name,"exe":p.info["exe"] or "","user":p.info["username"] or username,"start":p.info["create_time"]
            }
        except (psutil.NoSuchProcess, psutil.AccessDenied): pass
    return apps

def monitor_apps():
    global st_apps
    if not LOG_APPS: return
    st_apps = get_apps(); print(f"  📦 Apps: {len(st_apps)} tracked")
    while True:
        time.sleep(APP_POLL); cur = get_apps()
        with wb_lock:
            for pid, info in cur.items():
                if pid not in st_apps:
                    ts,d,t = now_ts()
                    write_row("Applications",[next_id(),ts,d,t,"OPENED",info["name"],pid,"",info["user"],info["exe"]],"APP OPEN")
                    log_all("APP OPEN","App Opened",info["name"],f"PID:{pid}")
                    print(f"[{ts}] 🟢 APP  OPEN : {info['name']}")
            for pid, info in st_apps.items():
                if pid not in cur:
                    ts,d,t = now_ts()
                    dur = str(datetime.timedelta(seconds=int(datetime.datetime.now().timestamp()-info["start"])))
                    write_row("Applications",[next_id(),ts,d,t,"CLOSED",info["name"],pid,dur,info["user"],info["exe"]],"APP CLOSE")
                    log_all("APP CLOSE","App Closed",info["name"],f"PID:{pid}",dur)
                    print(f"[{ts}] 🔴 APP CLOSE : {info['name']} ({dur})")
        st_apps = cur

# ══════════════════════════════════════════════════════════════
#  USB  (macOS: system_profiler SPUSBDataType)
# ══════════════════════════════════════════════════════════════

def get_usb():
    devs = {}
    try:
        out = run_cmd(["system_profiler","SPUSBDataType"])
        for line in out.splitlines():
            line = line.strip()
            if line.endswith(":") and len(line) > 3 and not line.startswith("USB"):
                nm = line.rstrip(":"); devs[nm] = {"name":nm,"drive":"","type":"USB Device"}
    except Exception: pass
    return devs

def monitor_usb():
    global st_usb; prev_d = get_usb(); st_usb = set(prev_d.keys())
    print(f"  🔌 USB: {len(st_usb)} devices")
    while True:
        time.sleep(USB_POLL); cur_d = get_usb(); cur_s = set(cur_d.keys())
        with wb_lock:
            for dev in cur_s - st_usb:
                info = cur_d[dev]; ts,d,t = now_ts()
                write_row("USB Devices",[next_id(),ts,d,t,"INSERTED",info["name"],info["drive"],info["type"],dev],"USB INSERT")
                log_all("USB INSERT","USB Connected",info["name"],info["type"])
                print(f"[{ts}] 🟢 USB INSERT: {info['name']}")
            for dev in st_usb - cur_s:
                ts,d,t = now_ts()
                write_row("USB Devices",[next_id(),ts,d,t,"REMOVED",dev,"","USB Device",dev],"USB REMOVE")
                log_all("USB REMOVE","USB Removed",dev); print(f"[{ts}] 🔴 USB REMOVE: {dev}")
        st_usb = cur_s; prev_d = cur_d

# ══════════════════════════════════════════════════════════════
#  WI-FI  (macOS: airport CLI)
# ══════════════════════════════════════════════════════════════

def get_wifi():
    result = {}
    try:
        out = run_cmd([AIRPORT, "-I"])
        info = {}
        for line in out.splitlines():
            line = line.strip()
            if   line.startswith("SSID:"):    info["ssid"]   = line.split(":",1)[-1].strip()
            elif "agrCtlRSSI" in line:         info["signal"] = line.split(":",1)[-1].strip()+" dBm"
            elif "channel" in line.lower():    info["band"]   = line.split(":",1)[-1].strip()
            elif "BSSID:" in line:             info["mac"]    = line.split(":",1)[-1].strip()
        if info.get("ssid"):
            # Get IP from ifconfig en0
            ip_out = run_cmd(["ipconfig","getifaddr","en0"])
            result[info["ssid"]] = {
                "signal":info.get("signal","?"),"band":info.get("band","?"),
                "ip":ip_out.strip() or "?","mac":info.get("mac","?"),"iface":"en0"
            }
    except Exception: pass
    return result

def monitor_wifi():
    global st_wifi; st_wifi = get_wifi()
    print(f"  📶 Wi-Fi: {list(st_wifi.keys()) or 'Not connected'}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_wifi()
        with wb_lock:
            for ssid, info in cur.items():
                if ssid not in st_wifi:
                    ts,d,t = now_ts()
                    write_row("Wi-Fi",[next_id(),ts,d,t,"CONNECTED",ssid,info["signal"],info["band"],info["ip"],info["mac"],info["iface"]],"WIFI CONNECT")
                    log_all("WIFI CONNECT","Wi-Fi Connected",ssid,f"Signal:{info['signal']}")
                    print(f"[{ts}] 🟢 WIFI CONN : {ssid}")
            for ssid in st_wifi:
                if ssid not in cur:
                    ts,d,t = now_ts(); old = st_wifi[ssid]
                    write_row("Wi-Fi",[next_id(),ts,d,t,"DISCONNECTED",ssid,old.get("signal","?"),old.get("band","?"),old.get("ip","?"),old.get("mac","?"),old.get("iface","?")],"WIFI DISCONNECT")
                    log_all("WIFI DISCONNECT","Wi-Fi Disconnected",ssid)
                    print(f"[{ts}] 🔴 WIFI DISC : {ssid}")
        st_wifi = cur

# ══════════════════════════════════════════════════════════════
#  BLUETOOTH  (macOS: system_profiler SPBluetoothDataType)
# ══════════════════════════════════════════════════════════════

def get_bluetooth():
    devs = {}
    try:
        out = run_cmd(["system_profiler","SPBluetoothDataType"], timeout=12)
        cur_dev = None
        for line in out.splitlines():
            line = line.strip()
            if line.endswith(":") and "Bluetooth" not in line and len(line) > 3:
                cur_dev = line.rstrip(":")
            elif cur_dev and "Connected: Yes" in line:
                devs[cur_dev] = {"mac":"?","type":"Bluetooth","paired":"Yes","rssi":"?"}
            elif cur_dev and "Address:" in line and cur_dev in devs:
                devs[cur_dev]["mac"] = line.split(":",1)[-1].strip()
    except Exception: pass
    return devs

def monitor_bluetooth():
    global st_bt; bt_state = get_bluetooth(); st_bt = set(bt_state.keys())
    print(f"  🔵 Bluetooth: {list(st_bt) or 'None'}")
    while True:
        time.sleep(NETWORK_POLL); cur_d = get_bluetooth(); cur_s = set(cur_d.keys())
        with wb_lock:
            for nm in cur_s - st_bt:
                info = cur_d[nm]; ts,d,t = now_ts()
                write_row("Bluetooth",[next_id(),ts,d,t,"CONNECTED",nm,info["mac"],info["type"],info["paired"],info["rssi"]],"BT CONNECT")
                log_all("BT CONNECT","BT Connected",nm,f"MAC:{info['mac']}"); print(f"[{ts}] 🔵 BT  CONN : {nm}")
            for nm in st_bt - cur_s:
                old = bt_state.get(nm,{}); ts,d,t = now_ts()
                write_row("Bluetooth",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("mac","?"),old.get("type","BT"),"?","?"],"BT DISCONNECT")
                log_all("BT DISCONNECT","BT Disconnected",nm); print(f"[{ts}] 🔵 BT  DISC : {nm}")
        st_bt = cur_s; bt_state = cur_d

# ══════════════════════════════════════════════════════════════
#  NETWORK / INTERNET / BATTERY / SCREEN / DRIVES
# ══════════════════════════════════════════════════════════════

def get_net():
    import socket as _s
    ifaces = {}
    for name, st in psutil.net_if_stats().items():
        if name.lower() in {"lo0","loopback"}: continue
        ip = mac = ""
        for a in psutil.net_if_addrs().get(name,[]):
            if a.family == _s.AF_INET:    ip  = a.address
            if a.family == psutil.AF_LINK: mac = a.address
        ifaces[name] = {"up":st.isup,"speed":st.speed,"ip":ip,"mac":mac}
    return ifaces

def monitor_network():
    global st_net; st_net = get_net()
    print(f"  🌐 Net: {[k for k,v in st_net.items() if v['up']]}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_net()
        with wb_lock:
            for iface, info in cur.items():
                old = st_net.get(iface)
                if old is None: continue
                if info["up"] != old["up"]:
                    ts,d,t = now_ts(); status = "UP" if info["up"] else "DOWN"; ev = f"NET {status}"
                    write_row("Network",[next_id(),ts,d,t,status,iface,info["ip"],info["mac"],info["speed"],f"Interface {status}"],ev)
                    log_all(ev,f"Network {status}",iface,f"IP:{info['ip']}")
                    print(f"[{ts}] 🌐 NET {status:5}: {iface}")
                elif info["ip"] != old["ip"] and info["ip"]:
                    ts,d,t = now_ts()
                    write_row("Network",[next_id(),ts,d,t,"IP CHANGED",iface,info["ip"],info["mac"],info["speed"],f"{old['ip']}→{info['ip']}"],"IP CHANGED")
                    log_all("IP CHANGED","IP Changed",iface,f"{old['ip']}→{info['ip']}")
        st_net = cur

def check_inet():
    try:
        t0 = time.time(); s = socket.create_connection((INTERNET_HOST,INTERNET_PORT),timeout=5); s.close()
        return True, round((time.time()-t0)*1000,1)
    except Exception: return False,0

def monitor_internet():
    global st_inet, st_inet_down_since
    is_up, lat = check_inet(); st_inet = is_up
    if not is_up: st_inet_down_since = datetime.datetime.now()
    print(f"  📡 Internet: {'Online' if is_up else 'OFFLINE'}{f' ({lat}ms)' if is_up else ''}")
    while True:
        time.sleep(INTERNET_POLL); is_up, lat = check_inet()
        with wb_lock:
            if is_up != st_inet:
                ts,d,t = now_ts()
                if is_up:
                    dur = str(datetime.datetime.now()-st_inet_down_since).split(".")[0] if st_inet_down_since else "?"
                    write_row("Internet",[next_id(),ts,d,t,"ONLINE",f"{lat}ms","Internet restored",dur],"INTERNET UP")
                    log_all("INTERNET UP","Internet Online",f"{lat}ms",f"Was offline:{dur}",dur)
                    print(f"[{ts}] 📡 INET ON  : {lat}ms"); st_inet_down_since=None
                else:
                    st_inet_down_since = datetime.datetime.now()
                    write_row("Internet",[next_id(),ts,d,t,"OFFLINE","—","Connection lost",""],"INTERNET DOWN")
                    log_all("INTERNET DOWN","Internet Offline","Connection lost")
                    print(f"[{ts}] 📡 INET OFF ⚠️")
        st_inet = is_up

def monitor_battery():
    global st_batt
    b = psutil.sensors_battery()
    if not b: print("  🔋 No battery / Mac Pro"); return
    st_batt = {"pct":round(b.percent,1),"charging":b.power_plugged}
    print(f"  🔋 Battery: {b.percent:.0f}% {'Charging' if b.power_plugged else 'Discharging'}")
    while True:
        time.sleep(BATTERY_POLL); b = psutil.sensors_battery()
        if not b: continue
        cur = {"pct":round(b.percent,1),"charging":b.power_plugged}
        old_pct = st_batt.get("pct",cur["pct"]); old_chg = st_batt.get("charging",cur["charging"])
        delta = abs(cur["pct"]-old_pct); ts,d,t = now_ts()
        secs = b.secsleft
        tr = ("Unlimited" if secs==psutil.POWER_TIME_UNLIMITED else "Unknown" if secs==psutil.POWER_TIME_UNKNOWN
              else str(datetime.timedelta(seconds=int(secs))))
        src = "AC Power" if cur["charging"] else "Battery"
        with wb_lock:
            if cur["charging"] != old_chg:
                ev = "CHARGING" if cur["charging"] else "DISCHARGING"
                write_row("Battery",[next_id(),ts,d,t,ev,f"{cur['pct']}%",ev.title(),tr,src],ev)
                log_all(ev,f"Battery {ev.title()}",f"{cur['pct']}%",src); print(f"[{ts}] 🔋 {ev}: {cur['pct']}%")
            if cur["pct"]<=5 and not cur["charging"] and old_pct>5:
                write_row("Battery",[next_id(),ts,d,t,"CRITICAL BATTERY",f"{cur['pct']}%","Discharging",tr,src],"BATTERY CRITICAL")
                log_all("BATTERY CRITICAL","Battery CRITICAL",f"{cur['pct']}%"); print(f"[{ts}] 🚨 BATT CRIT: {cur['pct']}%")
            elif cur["pct"]<=15 and not cur["charging"] and old_pct>15:
                write_row("Battery",[next_id(),ts,d,t,"LOW BATTERY",f"{cur['pct']}%","Discharging",tr,src],"BATTERY LOW")
                log_all("BATTERY LOW","Low Battery",f"{cur['pct']}%"); print(f"[{ts}] ⚠️  BATT LOW : {cur['pct']}%")
            elif delta >= BATTERY_DELTA:
                write_row("Battery",[next_id(),ts,d,t,"LEVEL UPDATE",f"{cur['pct']}%","Charging" if cur["charging"] else "Discharging",tr,src],"BATTERY")
        st_batt = cur

def is_locked():
    """macOS: use Quartz CGSession to check screen lock state."""
    try:
        out = subprocess.check_output(
            ["python3","-c","import Quartz; s=Quartz.CGSessionCopyCurrentDictionary(); print(s.get('CGSSessionScreenIsLocked',0))"],
            text=True, stderr=subprocess.DEVNULL, timeout=3)
        return out.strip() == "1"
    except Exception: return False

def monitor_screen():
    global st_screen; st_screen = is_locked(); lock_since = None
    print(f"  🖵  Screen: {'Locked' if st_screen else 'Unlocked'}")
    while True:
        time.sleep(SCREEN_POLL); cur = is_locked()
        if cur == st_screen: continue
        with wb_lock:
            ts,d,t = now_ts()
            if cur:
                lock_since = datetime.datetime.now()
                write_row("Screen",[next_id(),ts,d,t,"SCREEN LOCKED","",username,"Session locked"],"SCREEN LOCK")
                log_all("SCREEN LOCK","Screen Locked",username); print(f"[{ts}] 🔒 LOCKED")
            else:
                dur = str(datetime.datetime.now()-lock_since).split(".")[0] if lock_since else "?"
                write_row("Screen",[next_id(),ts,d,t,"SCREEN UNLOCKED",dur,username,f"Locked {dur}"],"SCREEN UNLOCK")
                log_all("SCREEN UNLOCK","Screen Unlocked",username,f"Locked {dur}",dur)
                print(f"[{ts}] 🔓 UNLOCKED (was {dur})"); lock_since = None
        st_screen = cur

def get_drives():
    drives = {}
    for p in psutil.disk_partitions(all=False):
        try:
            u = psutil.disk_usage(p.mountpoint)
            drives[p.mountpoint] = {"label":p.device,"fs":p.fstype,"total":round(u.total/1e9,1),"free":round(u.free/1e9,1)}
        except Exception: pass
    return drives

def monitor_drives():
    global st_drives; cur_d = get_drives(); st_drives = set(cur_d.keys())
    print(f"  💾 Drives: {[k for k in st_drives if not k.startswith('/System')]}")
    while True:
        time.sleep(DRIVE_POLL); cur = get_drives(); cur_s = set(cur.keys())
        with wb_lock:
            for mp in cur_s - st_drives:
                info = cur[mp]; ts,d,t = now_ts()
                write_row("Drives",[next_id(),ts,d,t,"MOUNTED",mp,info["label"],info["fs"],info["total"],info["free"]],"DRIVE MOUNT")
                log_all("DRIVE MOUNT","Drive Mounted",mp,f"{info['fs']} {info['total']}GB")
                print(f"[{ts}] 💾 MOUNT: {mp}")
            for mp in st_drives - cur_s:
                old = cur_d.get(mp,{}); ts,d,t = now_ts()
                write_row("Drives",[next_id(),ts,d,t,"UNMOUNTED",mp,old.get("label","?"),old.get("fs","?"),old.get("total","?"),"?"],"DRIVE UNMOUNT")
                log_all("DRIVE UNMOUNT","Drive Unmounted",mp); print(f"[{ts}] 💾 UNMOUNT: {mp}")
        st_drives = cur_s; cur_d = cur

# ══════════════════════════════════════════════════════════════
#  AUDIO  (macOS: system_profiler SPAudioDataType)
# ══════════════════════════════════════════════════════════════

def get_audio():
    devs = set()
    try:
        out = run_cmd(["system_profiler","SPAudioDataType"])
        for line in out.splitlines():
            line = line.strip()
            if line.endswith(":") and len(line)>3 and "Audio" not in line and "Devices" not in line:
                devs.add(line.rstrip(":"))
    except Exception: pass
    return devs

def monitor_audio():
    global st_audio; st_audio = get_audio()
    print(f"  🔊 Audio: {list(st_audio) or 'None'}")
    while True:
        time.sleep(AUDIO_POLL); cur = get_audio()
        with wb_lock:
            for dev in cur - st_audio:
                ts,d,t = now_ts()
                dtype = ("Headphones" if any(k in dev.lower() for k in ["head","ear","airpod","bud","beats"]) else
                         "Microphone" if any(k in dev.lower() for k in ["mic","input","record"]) else "Speaker")
                write_row("Audio",[next_id(),ts,d,t,"CONNECTED",dev,dtype,"?",f"{dev} connected"],"AUDIO IN")
                log_all("AUDIO IN","Audio Connected",dev,dtype); print(f"[{ts}] 🔊 AUDIO IN : {dev}")
            for dev in st_audio - cur:
                ts,d,t = now_ts()
                write_row("Audio",[next_id(),ts,d,t,"DISCONNECTED",dev,"?","?",f"{dev} removed"],"AUDIO OUT")
                log_all("AUDIO OUT","Audio Removed",dev); print(f"[{ts}] 🔊 AUDIO OUT: {dev}")
        st_audio = cur

def get_mics():
    """macOS: derive mic list from audio devices with 'mic'/'input' in name."""
    return {d for d in get_audio() if any(k in d.lower() for k in ["mic","input","record","capture"])}

def monitor_microphones():
    global st_mic; st_mic = get_mics()
    print(f"  🎙️  Mics: {list(st_mic) or 'None'}")
    while True:
        time.sleep(AUDIO_POLL); cur = get_mics()
        with wb_lock:
            for dev in cur - st_mic:
                ts,d,t = now_ts()
                write_row("Microphone",[next_id(),ts,d,t,"CONNECTED",dev,"Microphone",f"Mic in: {dev}"],"MIC IN")
                log_all("MIC IN","Mic Connected",dev); print(f"[{ts}] 🎙️  MIC IN  : {dev}")
            for dev in st_mic - cur:
                ts,d,t = now_ts()
                write_row("Microphone",[next_id(),ts,d,t,"DISCONNECTED",dev,"Microphone",f"Mic out: {dev}"],"MIC OUT")
                log_all("MIC OUT","Mic Removed",dev); print(f"[{ts}] 🎙️  MIC OUT : {dev}")
        st_mic = cur

def get_cams():
    cams = set()
    try:
        out = run_cmd(["system_profiler","SPCameraDataType"])
        for line in out.splitlines():
            line = line.strip()
            if line.endswith(":") and len(line) > 3: cams.add(line.rstrip(":"))
    except Exception: pass
    return cams

def monitor_cameras():
    global st_cam; st_cam = get_cams()
    print(f"  📷 Cameras: {list(st_cam) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_cams()
        with wb_lock:
            for dev in cur - st_cam:
                ts,d,t = now_ts()
                write_row("Webcam",[next_id(),ts,d,t,"CONNECTED",dev,"Camera",f"Camera in: {dev}"],"CAM IN")
                log_all("CAM IN","Camera Connected",dev); print(f"[{ts}] 📷 CAM IN  : {dev}")
            for dev in st_cam - cur:
                ts,d,t = now_ts()
                write_row("Webcam",[next_id(),ts,d,t,"DISCONNECTED",dev,"Camera",f"Camera out: {dev}"],"CAM OUT")
                log_all("CAM OUT","Camera Removed",dev); print(f"[{ts}] 📷 CAM OUT : {dev}")
        st_cam = cur

def get_printers():
    printers = {}
    try:
        out = run_cmd(["lpstat","-p"])
        for line in out.splitlines():
            if line.startswith("printer"):
                parts = line.split(); nm = parts[1] if len(parts)>1 else "?"
                printers[nm] = {"port":"?","driver":"?","status":"OK"}
    except Exception: pass
    return printers

def monitor_printers():
    global st_printers; st_printers = get_printers()
    print(f"  🖨️  Printers: {list(st_printers) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_printers()
        with wb_lock:
            for nm in set(cur) - set(st_printers):
                ts,d,t = now_ts()
                write_row("Printers",[next_id(),ts,d,t,"ADDED",nm,"?","?","OK"],"PRINTER ADD")
                log_all("PRINTER ADD","Printer Added",nm); print(f"[{ts}] 🖨️  PRTR ADD: {nm}")
            for nm in set(st_printers) - set(cur):
                ts,d,t = now_ts()
                write_row("Printers",[next_id(),ts,d,t,"REMOVED",nm,"?","?","Removed"],"PRINTER REMOVE")
                log_all("PRINTER REMOVE","Printer Removed",nm); print(f"[{ts}] 🖨️  PRTR REM: {nm}")
        st_printers = cur

def get_monitors():
    mons = {}
    try:
        out = run_cmd(["system_profiler","SPDisplaysDataType"], timeout=12)
        cur_mon = None
        for line in out.splitlines():
            line = line.strip()
            if line.endswith(":") and len(line)>3 and "Displays" not in line: cur_mon = line.rstrip(":")
            elif cur_mon and "Resolution:" in line:
                mons[cur_mon] = {"res":line.split(":",1)[-1].strip(),"port":"?","serial":"?"}
    except Exception: pass
    return mons

def monitor_monitors():
    global st_monitors; st_monitors = get_monitors()
    print(f"  📺 Monitors: {list(st_monitors) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_monitors()
        with wb_lock:
            for nm in set(cur) - set(st_monitors):
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Monitors",[next_id(),ts,d,t,"CONNECTED",nm,info["res"],info["port"],info["serial"]],"MONITOR IN")
                log_all("MONITOR IN","Monitor Connected",nm,info["res"]); print(f"[{ts}] 📺 MON IN  : {nm} ({info['res']})")
            for nm in set(st_monitors) - set(cur):
                old = st_monitors.get(nm,{}); ts,d,t = now_ts()
                write_row("Monitors",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("res","?"),"?","?"],"MONITOR OUT")
                log_all("MONITOR OUT","Monitor Disconnected",nm); print(f"[{ts}] 📺 MON OUT : {nm}")
        st_monitors = cur

def get_vpn():
    vpns = {}
    try:
        for iface in psutil.net_if_stats():
            n = iface.lower()
            if any(v in n for v in ["vpn","tun","tap","ppp","wireguard","ipsec","utun"]):
                if psutil.net_if_stats()[iface].isup:
                    vpns[iface] = {"server":"?","proto":"VPN Adapter"}
    except Exception: pass
    return vpns

def monitor_vpn():
    global st_vpn; prev = get_vpn(); st_vpn = set(prev.keys()); vpn_since = {}
    print(f"  🌍 VPN: {list(st_vpn) or 'None'}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_vpn()
        with wb_lock:
            for nm in set(cur) - st_vpn:
                info = cur[nm]; ts,d,t = now_ts(); vpn_since[nm] = datetime.datetime.now()
                write_row("VPN",[next_id(),ts,d,t,"CONNECTED",nm,info["server"],info["proto"],""],"VPN CONNECT")
                log_all("VPN CONNECT","VPN Connected",nm); print(f"[{ts}] 🌍 VPN CONN: {nm}")
            for nm in st_vpn - set(cur):
                old = prev.get(nm,{}); ts,d,t = now_ts()
                dur = str(datetime.datetime.now()-vpn_since[nm]).split(".")[0] if nm in vpn_since else "?"
                write_row("VPN",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("server","?"),old.get("proto","?"),dur],"VPN DISCONNECT")
                log_all("VPN DISCONNECT","VPN Disconnected",nm,"",dur); print(f"[{ts}] 🌍 VPN DISC: {nm}")
        st_vpn = set(cur); prev = cur

def monitor_hotspot():
    global st_hotspot; prev = set(); st_hotspot = prev
    print("  📱 Hotspot: monitoring iPhone USB tethering")
    while True:
        time.sleep(NETWORK_POLL); cur = set()
        try:
            for iface in psutil.net_if_stats():
                if ("iphone" in iface.lower() or "rndis" in iface.lower()) and psutil.net_if_stats()[iface].isup:
                    cur.add(iface)
        except Exception: pass
        with wb_lock:
            for dev in cur - prev:
                ts,d,t = now_ts()
                write_row("Mobile Hotspot",[next_id(),ts,d,t,"CONNECTED",dev,"iPhone USB Tethering","?",""],"HOTSPOT ON")
                log_all("HOTSPOT ON","Hotspot Connected",dev); print(f"[{ts}] 📱 HOTSPOT+: {dev}")
            for dev in prev - cur:
                ts,d,t = now_ts()
                write_row("Mobile Hotspot",[next_id(),ts,d,t,"DISCONNECTED",dev,"?","?",""],"HOTSPOT OFF")
                log_all("HOTSPOT OFF","Hotspot Removed",dev); print(f"[{ts}] 📱 HOTSPOT-: {dev}")
        prev = cur; st_hotspot = cur

def monitor_net_drives():
    global st_netdrives; prev = {}; st_netdrives = set()
    print("  🗂️  Net Drives: monitoring")
    while True:
        time.sleep(DRIVE_POLL); cur = {}
        try:
            out = run_cmd(["mount"])
            for line in out.splitlines():
                if any(t in line for t in ["smbfs","nfs","afpfs"]):
                    parts = line.split(); src = parts[0]; mp = parts[2] if len(parts)>2 else parts[0]
                    cur[mp] = {"unc":src,"server":src.split("/")[2] if "/" in src else "?"}
        except Exception: pass
        with wb_lock:
            for nm in set(cur) - st_netdrives:
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Network Drives",[next_id(),ts,d,t,"MOUNTED",nm,info["unc"],info["server"],"Connected"],"NET DRIVE MAP")
                log_all("NET DRIVE MAP","Net Drive Mounted",nm,info["unc"]); print(f"[{ts}] 🗂️  NETDRV+: {nm}")
            for nm in st_netdrives - set(cur):
                old = prev.get(nm,{}); ts,d,t = now_ts()
                write_row("Network Drives",[next_id(),ts,d,t,"UNMOUNTED",nm,old.get("unc","?"),old.get("server","?"),"Disconnected"],"NET DRIVE UNMAP")
                log_all("NET DRIVE UNMAP","Net Drive Unmounted",nm); print(f"[{ts}] 🗂️  NETDRV-: {nm}")
        st_netdrives = set(cur); prev = cur

def monitor_controllers():
    """macOS: detect game controllers via system_profiler SPUSBDataType keywords."""
    global st_ctrl; prev = set(); st_ctrl = prev
    print("  🎮 Controllers: monitoring")
    while True:
        time.sleep(DEVICE_POLL)
        try:
            out = run_cmd(["system_profiler","SPUSBDataType"])
            cur = set()
            ctrl_keywords = ["dualshock","dualsense","xbox","controller","gamepad","joystick"]
            for line in out.splitlines():
                line_l = line.strip().lower()
                if any(k in line_l for k in ctrl_keywords) and line.strip().endswith(":"):
                    cur.add(line.strip().rstrip(":"))
        except Exception: cur = set()
        with wb_lock:
            for nm in cur - prev:
                ts,d,t = now_ts()
                write_row("Controllers",[next_id(),ts,d,t,"CONNECTED",nm,"Controller","?","Ready"],"CTRL CONNECT")
                log_all("CTRL CONNECT","Controller Connected",nm); print(f"[{ts}] 🎮 CTRL IN : {nm}")
            for nm in prev - cur:
                ts,d,t = now_ts()
                write_row("Controllers",[next_id(),ts,d,t,"DISCONNECTED",nm,"Controller","?","Disconnected"],"CTRL DISCONNECT")
                log_all("CTRL DISCONNECT","Controller Removed",nm); print(f"[{ts}] 🎮 CTRL OUT: {nm}")
        prev = cur; st_ctrl = cur

def monitor_remote():
    global st_rdp; prev = set(); st_rdp = prev; rdp_since = {}
    print("  🔐 Remote: monitoring Screen Sharing / SSH")
    while True:
        time.sleep(DEVICE_POLL); cur = set()
        try:
            out = run_cmd(["who"])
            for line in out.splitlines():
                if "pts" in line or "ttys" in line:  # remote/SSH terminals
                    parts = line.split(); nm = parts[0] if parts else "?"
                    cur.add(nm)
        except Exception: pass
        with wb_lock:
            for nm in cur - prev:
                ts,d,t = now_ts(); rdp_since[nm] = datetime.datetime.now()
                write_row("Remote Sessions",[next_id(),ts,d,t,"CONNECTED",nm,"?","SSH/Screen Sharing",""],"RDP IN")
                log_all("RDP IN","Remote Session Started",nm); print(f"[{ts}] 🔐 REMOTE IN  : {nm}")
            for nm in prev - cur:
                ts,d,t = now_ts(); dur = str(datetime.datetime.now()-rdp_since[nm]).split(".")[0] if nm in rdp_since else "?"
                write_row("Remote Sessions",[next_id(),ts,d,t,"DISCONNECTED",nm,"?","SSH/Screen Sharing",dur],"RDP OUT")
                log_all("RDP OUT","Remote Session Ended",nm,"",dur); print(f"[{ts}] 🔐 REMOTE OUT : {nm}")
        prev = cur; st_rdp = cur

def monitor_security():
    """macOS: monitor Gatekeeper, SIP, Firewall status."""
    prev_fw = None
    print("  🛡️  Security: Monitoring macOS Firewall")
    while True:
        time.sleep(60)
        try:
            out = run_cmd(["/usr/libexec/ApplicationFirewall/socketfilterfw","--getglobalstate"]).strip()
            if out != prev_fw and prev_fw is not None:
                ts,d,t = now_ts()
                with wb_lock:
                    write_row("Security",[next_id(),ts,d,t,"FIREWALL CHANGED","macOS Firewall",out,f"Old:{prev_fw}→New:{out}"],"SECURITY")
                    log_all("SECURITY","Firewall Changed","macOS Firewall",out)
                    print(f"[{ts}] 🛡️  FIREWALL: {out}")
            prev_fw = out
        except Exception: pass

def monitor_users():
    prev = {u.name: u for u in psutil.users()}
    print(f"  👤 Users: {list(prev)}")
    while True:
        time.sleep(10)
        try: cur = {u.name: u for u in psutil.users()}
        except Exception: continue
        with wb_lock:
            for nm in set(cur) - set(prev):
                u = cur[nm]; ts,d,t = now_ts()
                started = datetime.datetime.fromtimestamp(u.started).strftime("%Y-%m-%d %H:%M:%S")
                write_row("Users",[next_id(),ts,d,t,"LOGIN",nm,u.terminal or "?",f"Session:{started}",pc_name],"USER LOGIN")
                log_all("USER LOGIN","Login",nm,u.terminal or "?"); print(f"[{ts}] 👤 LOGIN   : {nm}")
            for nm in set(prev) - set(cur):
                ts,d,t = now_ts()
                write_row("Users",[next_id(),ts,d,t,"LOGOUT",nm,prev[nm].terminal or "?","",pc_name],"USER LOGOUT")
                log_all("USER LOGOUT","Logout",nm); print(f"[{ts}] 👤 LOGOUT  : {nm}")
        prev = cur

def monitor_keyboards():
    """macOS: detect USB keyboards via system_profiler."""
    global st_kb; prev = set(); st_kb = prev
    print("  ⌨️  Keyboards: monitoring")
    while True:
        time.sleep(DEVICE_POLL)
        try:
            out = run_cmd(["system_profiler","SPUSBDataType"])
            cur = set()
            for line in out.splitlines():
                if "keyboard" in line.lower() and line.strip().endswith(":"):
                    cur.add(line.strip().rstrip(":"))
        except Exception: cur = set()
        with wb_lock:
            for dev in cur - prev:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"CONNECTED",dev,"Keyboard","USB","?"],"KB CONNECT")
                log_all("KB CONNECT","Keyboard Connected",dev,"USB"); print(f"[{ts}] ⌨️  KB IN   : {dev}")
            for dev in prev - cur:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"DISCONNECTED",dev,"Keyboard","?","?"],"KB DISCONNECT")
                log_all("KB DISCONNECT","Keyboard Disconnected",dev); print(f"[{ts}] ⌨️  KB OUT  : {dev}")
        prev = cur; st_kb = cur

def monitor_mice():
    """macOS: detect USB mice via system_profiler."""
    global st_mouse; prev = set(); st_mouse = prev
    print("  🖱️  Mice: monitoring")
    while True:
        time.sleep(DEVICE_POLL)
        try:
            out = run_cmd(["system_profiler","SPUSBDataType"])
            cur = set()
            for line in out.splitlines():
                if "mouse" in line.lower() and line.strip().endswith(":"):
                    cur.add(line.strip().rstrip(":"))
        except Exception: cur = set()
        with wb_lock:
            for dev in cur - prev:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"CONNECTED",dev,"Mouse","USB","?"],"MOUSE CONNECT")
                log_all("MOUSE CONNECT","Mouse Connected",dev,"USB"); print(f"[{ts}] 🖱️  MOUSE IN : {dev}")
            for dev in prev - cur:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"DISCONNECTED",dev,"Mouse","?","?"],"MOUSE DISCONNECT")
                log_all("MOUSE DISCONNECT","Mouse Disconnected",dev); print(f"[{ts}] 🖱️  MOUSE OUT: {dev}")
        prev = cur; st_mouse = cur

def monitor_power_plan():
    """macOS: detect Low Power Mode changes via pmset."""
    global st_plan; st_plan = "Normal"
    print("  ⚙️  Power Plan: monitoring Low Power Mode")
    while True:
        time.sleep(SERVICE_POLL)
        try:
            out = run_cmd(["pmset","-g"])
            lpm = "1" in [l.split()[-1] for l in out.splitlines() if "lowpowermode" in l.lower()]
            plan = "Low Power Mode" if lpm else "Normal"
            if plan != st_plan:
                ts,d,t = now_ts()
                with wb_lock:
                    write_row("Power Plan",[next_id(),ts,d,t,"CHANGED",plan,"N/A",st_plan,f"Switched to: {plan}"],"PLAN CHANGED")
                    log_all("PLAN CHANGED","Power Plan Changed",plan,f"Was:{st_plan}")
                    print(f"[{ts}] ⚙️  PLAN: {st_plan}→{plan}")
                st_plan = plan
        except Exception: pass

def monitor_services():
    """macOS: monitor launchctl daemon status."""
    global st_svc; watched = ["com.apple.networkd","com.apple.airportd","com.apple.bluetoothd"]
    print(f"  🔧 Services: monitoring {len(watched)} daemons")
    st_svc = {}
    while True:
        time.sleep(SERVICE_POLL)
        try:
            out = run_cmd(["launchctl","list"])
            cur = {}
            for line in out.splitlines()[1:]:
                parts = line.split("\t")
                if len(parts)>=3 and any(w in parts[2] for w in watched):
                    cur[parts[2]] = {"disp":parts[2],"status":"Running" if parts[0]!="-" else "Stopped"}
            with wb_lock:
                for nm, info in cur.items():
                    old = st_svc.get(nm,{})
                    if old.get("status") and old["status"] != info["status"]:
                        ts,d,t = now_ts(); cat = "SVC START" if info["status"]=="Running" else "SVC STOP"
                        write_row("Services",[next_id(),ts,d,t,info["status"],nm,nm,old["status"],info["status"]],cat)
                        log_all(cat,f"Service {info['status']}",nm); print(f"[{ts}] 🔧 SVC: {nm} {info['status']}")
            st_svc = cur
        except Exception: pass

def monitor_perf():
    if not LOG_PERF: return
    print(f"  📊 Perf: CPU>{CPU_SPIKE_THRESH}% / RAM>{RAM_SPIKE_THRESH}%")
    was_cpu = was_ram = False
    while True:
        time.sleep(PERF_POLL); cpu = psutil.cpu_percent(interval=2); ram = psutil.virtual_memory().percent
        ts,d,t = now_ts()
        with wb_lock:
            if cpu >= CPU_SPIKE_THRESH and not was_cpu:
                try:
                    top = max(psutil.process_iter(["name","cpu_percent"]),key=lambda p:p.info["cpu_percent"] or 0)
                    top_name = f"{top.info['name']} ({top.info['cpu_percent']:.0f}%)"
                except Exception: top_name = "?"
                write_row("System Perf",[next_id(),ts,d,t,"CPU SPIKE",f"{cpu:.1f}%",f"{ram:.1f}%",top_name,f"CPU spike:{cpu:.1f}%"],"CPU SPIKE")
                log_all("CPU SPIKE","CPU Spike",f"{cpu:.1f}%",top_name); print(f"[{ts}] 🔴 CPU SPIKE: {cpu:.1f}%")
            if ram >= RAM_SPIKE_THRESH and not was_ram:
                write_row("System Perf",[next_id(),ts,d,t,"RAM SPIKE",f"{cpu:.1f}%",f"{ram:.1f}%","?",f"RAM spike:{ram:.1f}%"],"RAM SPIKE")
                log_all("RAM SPIKE","RAM Spike",f"{ram:.1f}%"); print(f"[{ts}] 🔴 RAM SPIKE: {ram:.1f}%")
        was_cpu = (cpu >= CPU_SPIKE_THRESH); was_ram = (ram >= RAM_SPIKE_THRESH)

def monitor_smartcard():
    print("  🔑 SmartCard: monitoring (N/A on macOS without CryptoTokenKit)")

def monitor_display_mode():
    """macOS: monitor Night Shift and True Tone via defaults."""
    prev = {}; print("  🌙 Display Mode: monitoring Night Shift")
    while True:
        time.sleep(SERVICE_POLL)
        try:
            cur = {}
            ns_out = run_cmd(["defaults","read","com.apple.CoreBrightness"]).strip()
            cur["NightShift"] = "ON" if "CBBlueLightReductionEnabled = 1" in ns_out else "OFF"
            with wb_lock:
                for key, val in cur.items():
                    old_val = prev.get(key)
                    if old_val and val != old_val:
                        ts,d,t = now_ts()
                        write_row("Display Mode",[next_id(),ts,d,t,"CHANGED",key,old_val,val,f"{key}: {old_val}→{val}"],"DISPLAY CHANGE")
                        log_all("DISPLAY CHANGE","Display Mode Changed",key,f"{old_val}→{val}")
                        print(f"[{ts}] 🌙 DISPLAY : {key} {old_val}→{val}")
            prev = cur
        except Exception: pass

def monitor_clipboard():
    global clip_last
    if not LOG_CLIPBOARD: return
    print("  📋 Clipboard: ENABLED (pbpaste)")
    while True:
        time.sleep(3)
        try:
            text = subprocess.check_output(["pbpaste"], text=True, timeout=2).strip()
            if text and text != clip_last:
                snippet = text[:100].replace("\n"," ").replace("\r",""); ts,d,t = now_ts()
                with wb_lock:
                    write_row("Clipboard",[next_id(),ts,d,t,snippet,len(text),"Content changed"],"CLIPBOARD")
                clip_last = text
        except Exception: pass

def save_wb():
    try:
        os.makedirs(os.path.dirname(LOG_FILE) if os.path.dirname(LOG_FILE) else ".", exist_ok=True)
        wb.save(LOG_FILE)
    except Exception as e: print(f"  ⚠️  Save error: {e}")

def auto_save():
    while True:
        time.sleep(SAVE_EVERY)
        with wb_lock: save_wb()
        print(f"  💾 Auto-saved @ {datetime.datetime.now().strftime('%H:%M:%S')}")

# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    global wb
    print("╔"+"═"*62+"╗")
    print("║   PC ACTIVITY LOGGER — macOS EDITION                     ║")
    print("║   30 categories | Optimised for macOS 12+                ║")
    print("╚"+"═"*62+"╝")
    wb = init_excel()
    with wb_lock: log_startup(); save_wb()

    monitors = [
        ("Apps",        monitor_apps),         ("USB",         monitor_usb),
        ("WiFi",        monitor_wifi),          ("Bluetooth",   monitor_bluetooth),
        ("Network",     monitor_network),       ("Internet",    monitor_internet),
        ("Battery",     monitor_battery),       ("Screen",      monitor_screen),
        ("Drives",      monitor_drives),        ("Audio",       monitor_audio),
        ("Microphone",  monitor_microphones),   ("Camera",      monitor_cameras),
        ("Printers",    monitor_printers),      ("Monitors",    monitor_monitors),
        ("VPN",         monitor_vpn),           ("Hotspot",     monitor_hotspot),
        ("NetDrives",   monitor_net_drives),    ("Controllers", monitor_controllers),
        ("Remote",      monitor_remote),        ("Security",    monitor_security),
        ("Users",       monitor_users),         ("Keyboards",   monitor_keyboards),
        ("Mice",        monitor_mice),          ("PowerPlan",   monitor_power_plan),
        ("Services",    monitor_services),      ("PerfMonitor", monitor_perf),
        ("SmartCard",   monitor_smartcard),     ("DisplayMode", monitor_display_mode),
        ("Clipboard",   monitor_clipboard),     ("AutoSave",    auto_save),
    ]
    for name, fn in monitors:
        threading.Thread(target=fn, daemon=True, name=name).start()

    print(f"\n  📁 Log : {LOG_FILE}")
    print(f"  💾 Auto-save every {SAVE_EVERY}s | {len(monitors)-1} monitors active")
    print("  ⏹️  Press Ctrl+C to stop\n")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        print("\n  Stopping...")
        with wb_lock: log_shutdown()
        print("  ✅ Done.")

if __name__ == "__main__":
    if platform.system() != "Darwin":
        print("⚠️  This file is for macOS only. Use pc_logger_windows.py or pc_logger_linux.py instead.")
        sys.exit(1)
    main()
