"""
╔══════════════════════════════════════════════════════════════╗
║   PC ACTIVITY LOGGER — WINDOWS EDITION                       ║
║   Optimized exclusively for Windows 10 / 11                  ║
║   Requires: pip install psutil openpyxl pywin32              ║
╚══════════════════════════════════════════════════════════════╝

Windows-specific features:
  - WMI-based USB detection with serial IDs
  - PowerShell queries for Bluetooth, printers, monitors
  - Task Scheduler / Startup folder auto-run support
  - Windows Firewall, Defender, UAC change detection
  - VPN connection via Get-VpnConnection
  - Network drive mapping via Get-PSDrive
  - Power plan via powercfg
  - Clipboard via win32clipboard
  - Screen lock via ctypes User32
  - Smart card reader via PnpDevice
  - Game controller via HIDClass
  - RDP session detection via query session
"""

import os, sys, time, datetime, threading, platform, subprocess, socket, re
import psutil, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════

LOG_FILE = os.path.join(os.path.expanduser("~"), "OneDrive", "PC_ActivityLog.xlsx")
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Google Drive", "PC_ActivityLog.xlsx")
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Dropbox",      "PC_ActivityLog.xlsx")
# LOG_FILE = os.path.join(os.path.expanduser("~"), "Desktop",      "PC_ActivityLog.xlsx")

APP_POLL      = 5
USB_POLL      = 4
NETWORK_POLL  = 6
BATTERY_POLL  = 30
SCREEN_POLL   = 5
AUDIO_POLL    = 8
DRIVE_POLL    = 6
DEVICE_POLL   = 8
INTERNET_POLL = 15
PERF_POLL     = 20
SERVICE_POLL  = 30
SAVE_EVERY    = 30

BATTERY_DELTA    = 2
CPU_SPIKE_THRESH = 85
RAM_SPIKE_THRESH = 85

LOG_CLIPBOARD = False
LOG_APPS      = True
LOG_PERF      = True

WATCHED_SERVICES = [
    "WinDefend", "MpsSvc", "wuauserv", "BITS", "EventLog",
    "Spooler", "LanmanServer", "LanmanWorkstation", "Netlogon",
    "TermService", "RemoteRegistry", "W32Time", "Dnscache",
]

INTERNET_HOST = "8.8.8.8"
INTERNET_PORT = 53

# Windows background processes to exclude from app logging
IGNORED_PROCS = {
    "system idle process", "system", "registry", "smss.exe", "csrss.exe",
    "wininit.exe", "services.exe", "lsass.exe", "svchost.exe", "dwm.exe",
    "fontdrvhost.exe", "winlogon.exe", "spoolsv.exe", "taskhostw.exe",
    "runtimebroker.exe", "backgroundtaskhost.exe", "conhost.exe", "dllhost.exe",
    "sihost.exe", "ctfmon.exe", "searchindexer.exe", "wuauclt.exe", "msdtc.exe",
    "lsm.exe", "audiodg.exe", "sppsvc.exe", "wmpnetwk.exe", "dashost.exe",
    "explorer.exe", "securityhealthsystray.exe", "sgrmbroker.exe",
    "wlanext.exe", "nvdisplay.container.exe", "igfxem.exe",
}

# ══════════════════════════════════════════════════════════════
#  SHEET DEFINITIONS
# ══════════════════════════════════════════════════════════════

SHEETS = {
    "All Events":      ["#","Timestamp","Date","Time","Category","Event","Details","Extra","Duration","User","PC"],
    "Power":           ["#","Timestamp","Date","Time","Event","Details","Duration","PC"],
    "Applications":    ["#","Timestamp","Date","Time","Action","Application","PID","Duration","User","Path"],
    "USB Devices":     ["#","Timestamp","Date","Time","Action","Device Name","Drive","Type","Serial ID"],
    "Wi-Fi":           ["#","Timestamp","Date","Time","Action","SSID","Signal","Band","IP Address","MAC","Interface"],
    "Bluetooth":       ["#","Timestamp","Date","Time","Action","Device Name","MAC Address","Device Type","Paired","RSSI"],
    "Network":         ["#","Timestamp","Date","Time","Action","Interface","IP Address","MAC","Speed Mbps","Notes"],
    "Internet":        ["#","Timestamp","Date","Time","Status","Latency ms","Details","Duration Offline"],
    "Battery":         ["#","Timestamp","Date","Time","Event","Level %","Status","Time Remaining","Power Source"],
    "Screen":          ["#","Timestamp","Date","Time","Event","Duration Locked","User","Details"],
    "Drives":          ["#","Timestamp","Date","Time","Action","Drive","Label","Filesystem","Size GB","Free GB"],
    "Audio":           ["#","Timestamp","Date","Time","Action","Device Name","Category","Is Default","Details"],
    "Microphone":      ["#","Timestamp","Date","Time","Action","Device Name","Type","Details"],
    "Webcam":          ["#","Timestamp","Date","Time","Action","Device Name","Type","Details"],
    "Printers":        ["#","Timestamp","Date","Time","Action","Printer Name","Port","Driver","Status"],
    "Monitors":        ["#","Timestamp","Date","Time","Action","Monitor Name","Resolution","Port","Serial"],
    "VPN":             ["#","Timestamp","Date","Time","Action","VPN Name","Server","Protocol","Duration"],
    "Mobile Hotspot":  ["#","Timestamp","Date","Time","Action","Device Name","Type","IP","Details"],
    "Network Drives":  ["#","Timestamp","Date","Time","Action","Drive Letter","UNC Path","Server","Status"],
    "Controllers":     ["#","Timestamp","Date","Time","Action","Controller Name","Type","ID","Details"],
    "Remote Sessions": ["#","Timestamp","Date","Time","Action","Username","Client IP","Session Type","Duration"],
    "Security":        ["#","Timestamp","Date","Time","Event","Component","Status","Details"],
    "Users":           ["#","Timestamp","Date","Time","Event","Username","Session","Details","PC"],
    "Keyboard Mouse":  ["#","Timestamp","Date","Time","Action","Device Name","Type","Interface","Serial ID"],
    "Power Plan":      ["#","Timestamp","Date","Time","Action","Plan Name","Plan GUID","Previous Plan","Details"],
    "Services":        ["#","Timestamp","Date","Time","Action","Service Name","Display Name","Previous State","New State"],
    "System Perf":     ["#","Timestamp","Date","Time","Alert Type","CPU %","RAM %","Top Process","Details"],
    "Smart Card":      ["#","Timestamp","Date","Time","Action","Reader Name","Card Type","Details"],
    "Display Mode":    ["#","Timestamp","Date","Time","Action","Setting Changed","Old Value","New Value","Details"],
    "Clipboard":       ["#","Timestamp","Date","Time","Content Preview","Length","Details"],
    "Daily Summary":   ["Date","Power","Apps","USB","WiFi","BT","Internet","Printers","Monitors","VPN","Ctrl","Total"],
}

WIDTHS = {
    "All Events":      [5,22,12,10,16,18,42,18,12,10,14],
    "Power":           [5,22,12,10,22,50,14,16],
    "Applications":    [5,22,12,10,10,24,8,12,10,55],
    "USB Devices":     [5,22,12,10,12,30,10,18,28],
    "Wi-Fi":           [5,22,12,10,14,22,10,10,16,18,14],
    "Bluetooth":       [5,22,12,10,14,28,18,18,10,12],
    "Network":         [5,22,12,10,14,16,16,18,12,20],
    "Internet":        [5,22,12,10,14,12,30,16],
    "Battery":         [5,22,12,10,18,10,16,16,16],
    "Screen":          [5,22,12,10,16,16,14,30],
    "Drives":          [5,22,12,10,12,14,18,14,10,10],
    "Audio":           [5,22,12,10,14,35,16,12,30],
    "Microphone":      [5,22,12,10,14,35,16,30],
    "Webcam":          [5,22,12,10,14,35,16,30],
    "Printers":        [5,22,12,10,14,30,18,25,14],
    "Monitors":        [5,22,12,10,14,30,16,16,20],
    "VPN":             [5,22,12,10,14,22,20,14,14],
    "Mobile Hotspot":  [5,22,12,10,14,25,16,16,25],
    "Network Drives":  [5,22,12,10,12,14,35,20,14],
    "Controllers":     [5,22,12,10,14,28,16,12,25],
    "Remote Sessions": [5,22,12,10,14,18,18,14,14],
    "Security":        [5,22,12,10,18,18,14,40],
    "Users":           [5,22,12,10,14,16,12,35,14],
    "Keyboard Mouse":  [5,22,12,10,14,30,16,14,25],
    "Power Plan":      [5,22,12,10,14,25,35,25,25],
    "Services":        [5,22,12,10,14,20,30,16,16],
    "System Perf":     [5,22,12,10,16,10,10,25,35],
    "Smart Card":      [5,22,12,10,14,28,18,30],
    "Display Mode":    [5,22,12,10,14,22,18,18,30],
    "Clipboard":       [5,22,12,10,55,8,20],
    "Daily Summary":   [14,8,8,8,8,8,8,8,8,8,8,10],
}

CATS = {
    "POWER ON":"C6EFCE","POWER OFF":"FFCCCC","SLEEP":"FFF2CC","WAKE":"E2EFDA",
    "HIBERNATE":"FFE4B5","RESUME":"D5F5E3","APP OPEN":"DDEBF7","APP CLOSE":"FCE4D6",
    "USB INSERT":"E2EFDA","USB REMOVE":"FFF2CC","WIFI CONNECT":"D9EAD3","WIFI DISCONNECT":"FCE4D6",
    "BT CONNECT":"CFE2F3","BT DISCONNECT":"FCE4D6","NET UP":"D9EAD3","NET DOWN":"FFCCCC",
    "IP CHANGED":"FFF9C4","INTERNET UP":"D9EAD3","INTERNET DOWN":"FFCCCC",
    "BATTERY LOW":"FFCCCC","BATTERY CRITICAL":"FF9999","CHARGING":"C6EFCE",
    "DISCHARGING":"FFF2CC","BATTERY":"FFF9C4","SCREEN LOCK":"E8DAEF","SCREEN UNLOCK":"D5F5E3",
    "DRIVE MOUNT":"E2EFDA","DRIVE UNMOUNT":"FFF2CC","AUDIO IN":"E8F4FD","AUDIO OUT":"FCE4D6",
    "MIC IN":"E8F4FD","MIC OUT":"FCE4D6","CAM IN":"E2EFDA","CAM OUT":"FFF2CC",
    "PRINTER ADD":"E2EFDA","PRINTER REMOVE":"FFF2CC","MONITOR IN":"C6EFCE","MONITOR OUT":"FFCCCC",
    "VPN CONNECT":"D9EAD3","VPN DISCONNECT":"FFCCCC","HOTSPOT ON":"FFF9C4","HOTSPOT OFF":"FCE4D6",
    "NET DRIVE MAP":"E2EFDA","NET DRIVE UNMAP":"FFF2CC","CTRL CONNECT":"E2EFDA","CTRL DISCONNECT":"FFF2CC",
    "RDP IN":"CFE2F3","RDP OUT":"FCE4D6","SECURITY":"FFCCCC","SECURITY OK":"E2EFDA",
    "USER LOGIN":"C6EFCE","USER LOGOUT":"FFCCCC","KB CONNECT":"E2EFDA","KB DISCONNECT":"FFF2CC",
    "MOUSE CONNECT":"E2EFDA","MOUSE DISCONNECT":"FFF2CC","PLAN CHANGED":"FFF9C4",
    "SVC START":"D9EAD3","SVC STOP":"FFCCCC","CPU SPIKE":"FFCCCC","RAM SPIKE":"FFE4B5",
    "CARD IN":"E2EFDA","CARD OUT":"FFF2CC","DISPLAY CHANGE":"FFF9C4","CLIPBOARD":"F5F5F5",
}

HDR_FILL = PatternFill("solid", fgColor="1F3864")

# ══════════════════════════════════════════════════════════════
#  EXCEL HELPERS
# ══════════════════════════════════════════════════════════════

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_header_row(ws, columns):
    for i, title in enumerate(columns, 1):
        c = ws.cell(1, i, title)
        c.font      = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(bottom=Side(style="medium", color="FFFFFF"))
    ws.row_dimensions[1].height = 23

def style_data_row(ws, row_num, category=None):
    fill = (make_fill(CATS[category]) if category and category in CATS
            else make_fill("EEF2FF") if row_num % 2 == 0
            else make_fill("FFFFFF"))
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row_num, col)
        c.fill = fill
        c.font = Font(name="Segoe UI", size=9)
        c.alignment = Alignment(vertical="center")
        c.border = Border(bottom=Side(style="hair", color="D9D9D9"))

# ══════════════════════════════════════════════════════════════
#  GLOBAL STATE
# ══════════════════════════════════════════════════════════════

wb = None
wb_lock      = threading.Lock()
row_counters = {}
event_id     = 0

boot_time     = datetime.datetime.fromtimestamp(psutil.boot_time())
session_start = datetime.datetime.now()
pc_name       = socket.gethostname()
username      = os.getenv("USERNAME", "Unknown")

st_apps={}; st_usb=set(); st_wifi={}; st_bt=set(); st_net={}
st_batt={}; st_screen=None; st_drives=set(); st_audio=set()
st_mic=set(); st_cam=set(); st_printers={}; st_monitors={}
st_vpn=set(); st_hotspot=set(); st_netdrives=set(); st_ctrl=set()
st_rdp=set(); st_kb=set(); st_mouse=set(); st_svc={}; st_card=set()
st_plan=""; st_inet=None; st_inet_down_since=None; clip_last=""

def next_row(sheet):
    if sheet not in row_counters:
        row_counters[sheet] = max(wb[sheet].max_row, 1) + 1
    row_counters[sheet] += 1
    return row_counters[sheet] - 1

def next_id():
    global event_id; event_id += 1; return event_id

def write_row(sheet, data, category=None):
    row = next_row(sheet)
    ws  = wb[sheet]
    for col, val in enumerate(data, 1):
        ws.cell(row, col, val)
    style_data_row(ws, row, category)

def now_ts():
    dt = datetime.datetime.now()
    return dt.strftime("%Y-%m-%d %H:%M:%S"), dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")

def log_all(cat, etype, details, extra="", duration=""):
    ts, d, t = now_ts()
    write_row("All Events", [next_id(),ts,d,t,cat,etype,details,extra,duration,username,pc_name], cat)

def run_ps(cmd, timeout=10):
    """Run a PowerShell command, return stdout string."""
    try:
        return subprocess.check_output(
            ["powershell", "-Command", cmd],
            text=True, stderr=subprocess.DEVNULL, timeout=timeout
        )
    except Exception:
        return ""

# ══════════════════════════════════════════════════════════════
#  EXCEL INIT
# ══════════════════════════════════════════════════════════════

def init_excel():
    if os.path.exists(LOG_FILE):
        print(f"  📂 Existing log loaded: {LOG_FILE}")
        return openpyxl.load_workbook(LOG_FILE)
    wb2 = openpyxl.Workbook(); wb2.remove(wb2.active)
    info = wb2.create_sheet("INFO")
    info["B2"] = "PC ACTIVITY LOGGER — WINDOWS EDITION"
    info["B2"].font = Font(bold=True, size=15, color="1F3864", name="Segoe UI")
    for i, (lbl, val) in enumerate([
        ("PC Name:", pc_name), ("Started:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("User:", username), ("OS:", platform.platform()),
        ("IP:", socket.gethostbyname(socket.gethostname())), ("Python:", sys.version.split()[0])
    ]):
        info.cell(4+i, 2, lbl).font = Font(bold=True, color="595959", name="Segoe UI")
        info.cell(4+i, 3, val).font = Font(name="Segoe UI")
    info.column_dimensions["B"].width = 20
    info.column_dimensions["C"].width = 58
    for name, cols in SHEETS.items():
        ws = wb2.create_sheet(name)
        write_header_row(ws, cols)
        ws.freeze_panes = "A2"
        if name in WIDTHS:
            for ci, w in enumerate(WIDTHS[name], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
    os.makedirs(os.path.dirname(LOG_FILE) if os.path.dirname(LOG_FILE) else ".", exist_ok=True)
    wb2.save(LOG_FILE)
    return wb2

# ══════════════════════════════════════════════════════════════
#  POWER
# ══════════════════════════════════════════════════════════════

def log_startup():
    ts, d, t = now_ts()
    write_row("Power", [next_id(),ts,d,t,"POWER ON",
        f"Boot:{boot_time.strftime('%Y-%m-%d %H:%M:%S')} Logger:{ts}","",pc_name], "POWER ON")
    log_all("POWER ON","System Startup",f"Boot:{boot_time.strftime('%H:%M:%S')}",platform.platform())
    print(f"[{ts}] ✅ Windows Logger → {LOG_FILE}")

def log_shutdown():
    ts, d, t = now_ts()
    dur = str(datetime.datetime.now() - session_start).split(".")[0]
    write_row("Power", [next_id(),ts,d,t,"POWER OFF",f"Session:{dur}",dur,pc_name], "POWER OFF")
    log_all("POWER OFF","System Shutdown",f"Session:{dur}","",dur)
    save_wb(); print(f"[{ts}] 💾 Logger stopped — Session: {dur}")

# ══════════════════════════════════════════════════════════════
#  APPLICATIONS  (Windows: uses psutil with IGNORED_PROCS)
# ══════════════════════════════════════════════════════════════

def get_apps():
    apps = {}
    for p in psutil.process_iter(["pid","name","exe","username","create_time"]):
        try:
            name = p.info["name"] or ""
            if name.lower() in IGNORED_PROCS or not name.strip():
                continue
            apps[p.info["pid"]] = {
                "name": name, "exe": p.info["exe"] or "",
                "user": p.info["username"] or username,
                "start": p.info["create_time"]
            }
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass
    return apps

def monitor_apps():
    global st_apps
    if not LOG_APPS: return
    st_apps = get_apps()
    print(f"  📦 Apps: {len(st_apps)} tracked")
    while True:
        time.sleep(APP_POLL); cur = get_apps()
        with wb_lock:
            for pid, info in cur.items():
                if pid not in st_apps:
                    ts,d,t = now_ts()
                    write_row("Applications",[next_id(),ts,d,t,"OPENED",info["name"],pid,"",info["user"],info["exe"]],"APP OPEN")
                    log_all("APP OPEN","App Opened",info["name"],f"PID:{pid}")
                    print(f"[{ts}] 🟢 APP  OPEN : {info['name']}")
            for pid, info in st_apps.items():
                if pid not in cur:
                    ts,d,t = now_ts()
                    dur = str(datetime.timedelta(seconds=int(datetime.datetime.now().timestamp()-info["start"])))
                    write_row("Applications",[next_id(),ts,d,t,"CLOSED",info["name"],pid,dur,info["user"],info["exe"]],"APP CLOSE")
                    log_all("APP CLOSE","App Closed",info["name"],f"PID:{pid}",dur)
                    print(f"[{ts}] 🔴 APP CLOSE : {info['name']} ({dur})")
        st_apps = cur

# ══════════════════════════════════════════════════════════════
#  USB  (Windows: WMI Win32_USBHub + Win32_LogicalDisk)
# ══════════════════════════════════════════════════════════════

def get_usb():
    devs = {}
    try:
        import win32com.client
        wmi = win32com.client.GetObject("winmgmts:")
        for d in wmi.InstancesOf("Win32_USBHub"):
            devs[d.DeviceID] = {"name": d.Description or d.DeviceID, "drive":"", "type":"USB Hub"}
        for d in wmi.InstancesOf("Win32_LogicalDisk"):
            if d.DriveType == 2:
                devs[f"DRIVE_{d.DeviceID}"] = {"name": d.VolumeName or "USB Drive", "drive": d.DeviceID, "type":"Removable Drive"}
    except Exception:
        pass
    return devs

def monitor_usb():
    global st_usb
    st_usb = set(get_usb().keys())
    print(f"  🔌 USB: {len(st_usb)} devices")
    while True:
        time.sleep(USB_POLL); cur_d = get_usb(); cur_s = set(cur_d.keys())
        with wb_lock:
            for dev in cur_s - st_usb:
                info = cur_d[dev]; ts,d,t = now_ts()
                write_row("USB Devices",[next_id(),ts,d,t,"INSERTED",info["name"],info["drive"],info["type"],dev],"USB INSERT")
                log_all("USB INSERT","USB Connected",info["name"],info["type"])
                print(f"[{ts}] 🟢 USB INSERT: {info['name']}")
            for dev in st_usb - cur_s:
                ts,d,t = now_ts()
                write_row("USB Devices",[next_id(),ts,d,t,"REMOVED",dev,"","USB Device",dev],"USB REMOVE")
                log_all("USB REMOVE","USB Removed",dev)
                print(f"[{ts}] 🔴 USB REMOVE: {dev}")
        st_usb = cur_s

# ══════════════════════════════════════════════════════════════
#  WI-FI  (Windows: netsh wlan show interfaces)
# ══════════════════════════════════════════════════════════════

def get_wifi():
    result = {}
    try:
        out = subprocess.check_output(["netsh","wlan","show","interfaces"],
            text=True, stderr=subprocess.DEVNULL, timeout=8)
        info = {}
        for line in out.splitlines():
            line = line.strip()
            if   "SSID" in line and "BSSID" not in line: info["ssid"]   = line.split(":",1)[-1].strip()
            elif "Signal"       in line:                  info["signal"] = line.split(":",1)[-1].strip()
            elif "Radio type"   in line:                  info["band"]   = line.split(":",1)[-1].strip()
            elif "IPv4 Address" in line:                  info["ip"]     = line.split(":",1)[-1].strip()
            elif "Physical address" in line:              info["mac"]    = line.split(":",1)[-1].strip()
            elif line.startswith("Name"):                 info["iface"]  = line.split(":",1)[-1].strip()
        if info.get("ssid"):
            result[info["ssid"]] = {
                "signal": info.get("signal","?"), "band": info.get("band","?"),
                "ip":     info.get("ip","?"),     "mac":  info.get("mac","?"),
                "iface":  info.get("iface","?")
            }
    except Exception:
        pass
    return result

def monitor_wifi():
    global st_wifi
    st_wifi = get_wifi()
    print(f"  📶 Wi-Fi: {list(st_wifi.keys()) or 'Not connected'}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_wifi()
        with wb_lock:
            for ssid, info in cur.items():
                if ssid not in st_wifi:
                    ts,d,t = now_ts()
                    write_row("Wi-Fi",[next_id(),ts,d,t,"CONNECTED",ssid,info["signal"],info["band"],info["ip"],info["mac"],info["iface"]],"WIFI CONNECT")
                    log_all("WIFI CONNECT","Wi-Fi Connected",ssid,f"Signal:{info['signal']} IP:{info['ip']}")
                    print(f"[{ts}] 🟢 WIFI CONN : {ssid}")
            for ssid in st_wifi:
                if ssid not in cur:
                    ts,d,t = now_ts(); old = st_wifi[ssid]
                    write_row("Wi-Fi",[next_id(),ts,d,t,"DISCONNECTED",ssid,old.get("signal","?"),old.get("band","?"),old.get("ip","?"),old.get("mac","?"),old.get("iface","?")],"WIFI DISCONNECT")
                    log_all("WIFI DISCONNECT","Wi-Fi Disconnected",ssid)
                    print(f"[{ts}] 🔴 WIFI DISC : {ssid}")
        st_wifi = cur

# ══════════════════════════════════════════════════════════════
#  BLUETOOTH  (Windows: Get-PnpDevice -Class Bluetooth)
# ══════════════════════════════════════════════════════════════

def get_bluetooth():
    devs = {}
    try:
        out = run_ps("Get-PnpDevice -Class Bluetooth | Where-Object {$_.Status -eq 'OK'} | Select-Object FriendlyName,DeviceID | Format-List")
        name = None
        for line in out.splitlines():
            line = line.strip()
            if line.startswith("FriendlyName"):
                name = line.split(":",1)[-1].strip()
            elif line.startswith("DeviceID") and name:
                did = line.split(":",1)[-1].strip()
                m   = re.search(r"([0-9A-Fa-f]{12})", did.replace("_","").replace("-",""))
                mac = ":".join(m.group(0)[i:i+2] for i in range(0,12,2)) if m else "?"
                devs[name] = {"mac":mac,"type":"Bluetooth","paired":"Yes","rssi":"?"}
                name = None
    except Exception:
        pass
    return devs

def monitor_bluetooth():
    global st_bt
    bt_state = get_bluetooth(); st_bt = set(bt_state.keys())
    print(f"  🔵 Bluetooth: {list(st_bt) or 'None'}")
    while True:
        time.sleep(NETWORK_POLL); cur_d = get_bluetooth(); cur_s = set(cur_d.keys())
        with wb_lock:
            for nm in cur_s - st_bt:
                info = cur_d[nm]; ts,d,t = now_ts()
                write_row("Bluetooth",[next_id(),ts,d,t,"CONNECTED",nm,info["mac"],info["type"],info["paired"],info["rssi"]],"BT CONNECT")
                log_all("BT CONNECT","BT Connected",nm,f"MAC:{info['mac']}")
                print(f"[{ts}] 🔵 BT  CONN : {nm}")
            for nm in st_bt - cur_s:
                old = bt_state.get(nm,{}); ts,d,t = now_ts()
                write_row("Bluetooth",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("mac","?"),old.get("type","BT"),"?","?"],"BT DISCONNECT")
                log_all("BT DISCONNECT","BT Disconnected",nm)
                print(f"[{ts}] 🔵 BT  DISC : {nm}")
        st_bt = cur_s; bt_state = cur_d

# ══════════════════════════════════════════════════════════════
#  NETWORK / INTERNET / BATTERY / SCREEN / DRIVES
#  (same cross-platform logic — included from shared helpers)
# ══════════════════════════════════════════════════════════════

def get_net():
    import socket as _s
    ifaces = {}
    for name, st in psutil.net_if_stats().items():
        if name.lower() in {"lo","loopback"}: continue
        ip = mac = ""
        for a in psutil.net_if_addrs().get(name,[]):
            if a.family == _s.AF_INET:    ip  = a.address
            if a.family == psutil.AF_LINK: mac = a.address
        ifaces[name] = {"up":st.isup,"speed":st.speed,"ip":ip,"mac":mac}
    return ifaces

def monitor_network():
    global st_net; st_net = get_net()
    print(f"  🌐 Net: {[k for k,v in st_net.items() if v['up']]}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_net()
        with wb_lock:
            for iface, info in cur.items():
                old = st_net.get(iface)
                if old is None: continue
                if info["up"] != old["up"]:
                    ts,d,t = now_ts(); status = "UP" if info["up"] else "DOWN"
                    ev = f"NET {status}"
                    notes = ("Ethernet plugged in" if info["up"] and "eth" in iface.lower() else
                             "Ethernet unplugged"  if not info["up"] and "eth" in iface.lower() else
                             f"Interface {status}")
                    write_row("Network",[next_id(),ts,d,t,status,iface,info["ip"],info["mac"],info["speed"],notes],ev)
                    log_all(ev,f"Network {status}",iface,f"IP:{info['ip']}")
                    print(f"[{ts}] 🌐 NET {status:5}: {iface}")
                elif info["ip"] != old["ip"] and info["ip"]:
                    ts,d,t = now_ts()
                    write_row("Network",[next_id(),ts,d,t,"IP CHANGED",iface,info["ip"],info["mac"],info["speed"],f"{old['ip']}→{info['ip']}"],"IP CHANGED")
                    log_all("IP CHANGED","IP Changed",iface,f"{old['ip']}→{info['ip']}")
        st_net = cur

def check_inet():
    try:
        t0 = time.time()
        s  = socket.create_connection((INTERNET_HOST, INTERNET_PORT), timeout=5)
        s.close()
        return True, round((time.time()-t0)*1000,1)
    except Exception:
        return False, 0

def monitor_internet():
    global st_inet, st_inet_down_since
    is_up, lat = check_inet(); st_inet = is_up
    if not is_up: st_inet_down_since = datetime.datetime.now()
    print(f"  📡 Internet: {'Online' if is_up else 'OFFLINE'}{f' ({lat}ms)' if is_up else ''}")
    while True:
        time.sleep(INTERNET_POLL); is_up, lat = check_inet()
        with wb_lock:
            if is_up != st_inet:
                ts,d,t = now_ts()
                if is_up:
                    dur = str(datetime.datetime.now()-st_inet_down_since).split(".")[0] if st_inet_down_since else "?"
                    write_row("Internet",[next_id(),ts,d,t,"ONLINE",f"{lat}ms","Internet restored",dur],"INTERNET UP")
                    log_all("INTERNET UP","Internet Online",f"{lat}ms",f"Was offline:{dur}",dur)
                    print(f"[{ts}] 📡 INET ON  : {lat}ms"); st_inet_down_since = None
                else:
                    st_inet_down_since = datetime.datetime.now()
                    write_row("Internet",[next_id(),ts,d,t,"OFFLINE","—","Connection lost",""],"INTERNET DOWN")
                    log_all("INTERNET DOWN","Internet Offline","Connection lost")
                    print(f"[{ts}] 📡 INET OFF ⚠️")
        st_inet = is_up

def monitor_battery():
    global st_batt
    b = psutil.sensors_battery()
    if not b: print("  🔋 No battery (desktop)"); return
    st_batt = {"pct": round(b.percent,1), "charging": b.power_plugged}
    print(f"  🔋 Battery: {b.percent:.0f}% {'Charging' if b.power_plugged else 'Discharging'}")
    while True:
        time.sleep(BATTERY_POLL); b = psutil.sensors_battery()
        if not b: continue
        cur = {"pct": round(b.percent,1), "charging": b.power_plugged}
        old_pct = st_batt.get("pct", cur["pct"]); old_chg = st_batt.get("charging", cur["charging"])
        delta = abs(cur["pct"] - old_pct); ts,d,t = now_ts()
        secs = b.secsleft
        tr = ("Unlimited" if secs==psutil.POWER_TIME_UNLIMITED else
              "Unknown"   if secs==psutil.POWER_TIME_UNKNOWN   else
              str(datetime.timedelta(seconds=int(secs))))
        src = "AC Power" if cur["charging"] else "Battery"
        with wb_lock:
            if cur["charging"] != old_chg:
                ev = "CHARGING" if cur["charging"] else "DISCHARGING"
                write_row("Battery",[next_id(),ts,d,t,ev,f"{cur['pct']}%",ev.title(),tr,src],ev)
                log_all(ev,f"Battery {ev.title()}",f"{cur['pct']}%",src)
                print(f"[{ts}] 🔋 {ev}: {cur['pct']}%")
            if cur["pct"]<=5 and not cur["charging"] and old_pct>5:
                write_row("Battery",[next_id(),ts,d,t,"CRITICAL BATTERY",f"{cur['pct']}%","Discharging",tr,src],"BATTERY CRITICAL")
                log_all("BATTERY CRITICAL","Battery CRITICAL",f"{cur['pct']}%",f"Est:{tr}")
                print(f"[{ts}] 🚨 BATT CRIT: {cur['pct']}%")
            elif cur["pct"]<=15 and not cur["charging"] and old_pct>15:
                write_row("Battery",[next_id(),ts,d,t,"LOW BATTERY",f"{cur['pct']}%","Discharging",tr,src],"BATTERY LOW")
                log_all("BATTERY LOW","Low Battery",f"{cur['pct']}%",f"Est:{tr}")
                print(f"[{ts}] ⚠️  BATT LOW : {cur['pct']}%")
            elif delta >= BATTERY_DELTA:
                write_row("Battery",[next_id(),ts,d,t,"LEVEL UPDATE",f"{cur['pct']}%","Charging" if cur["charging"] else "Discharging",tr,src],"BATTERY")
        st_batt = cur

def is_locked():
    try:
        import ctypes
        return ctypes.windll.User32.GetForegroundWindow() == 0
    except Exception:
        return False

def monitor_screen():
    global st_screen; st_screen = is_locked(); lock_since = None
    print(f"  🖵  Screen: {'Locked' if st_screen else 'Unlocked'}")
    while True:
        time.sleep(SCREEN_POLL); cur = is_locked()
        if cur == st_screen: continue
        with wb_lock:
            ts,d,t = now_ts()
            if cur:
                lock_since = datetime.datetime.now()
                write_row("Screen",[next_id(),ts,d,t,"SCREEN LOCKED","",username,"Session locked"],"SCREEN LOCK")
                log_all("SCREEN LOCK","Screen Locked",username); print(f"[{ts}] 🔒 LOCKED")
            else:
                dur = str(datetime.datetime.now()-lock_since).split(".")[0] if lock_since else "?"
                write_row("Screen",[next_id(),ts,d,t,"SCREEN UNLOCKED",dur,username,f"Locked {dur}"],"SCREEN UNLOCK")
                log_all("SCREEN UNLOCK","Screen Unlocked",username,f"Locked {dur}",dur)
                print(f"[{ts}] 🔓 UNLOCKED (was {dur})"); lock_since = None
        st_screen = cur

def get_drives():
    drives = {}
    for p in psutil.disk_partitions(all=False):
        try:
            u = psutil.disk_usage(p.mountpoint)
            drives[p.mountpoint] = {"label":p.device,"fs":p.fstype,"total":round(u.total/1e9,1),"free":round(u.free/1e9,1)}
        except Exception:
            pass
    return drives

def monitor_drives():
    global st_drives
    cur_d = get_drives(); st_drives = set(cur_d.keys())
    print(f"  💾 Drives: {list(st_drives)}")
    while True:
        time.sleep(DRIVE_POLL); cur = get_drives(); cur_s = set(cur.keys())
        with wb_lock:
            for mp in cur_s - st_drives:
                info = cur[mp]; ts,d,t = now_ts()
                write_row("Drives",[next_id(),ts,d,t,"MOUNTED",mp,info["label"],info["fs"],info["total"],info["free"]],"DRIVE MOUNT")
                log_all("DRIVE MOUNT","Drive Mounted",mp,f"{info['fs']} {info['total']}GB")
                print(f"[{ts}] 💾 MOUNT: {mp}")
            for mp in st_drives - cur_s:
                old = cur_d.get(mp,{}); ts,d,t = now_ts()
                write_row("Drives",[next_id(),ts,d,t,"UNMOUNTED",mp,old.get("label","?"),old.get("fs","?"),old.get("total","?"),"?"],"DRIVE UNMOUNT")
                log_all("DRIVE UNMOUNT","Drive Unmounted",mp); print(f"[{ts}] 💾 UNMOUNT: {mp}")
        st_drives = cur_s; cur_d = cur

# ══════════════════════════════════════════════════════════════
#  AUDIO  (Windows: Get-PnpDevice -Class AudioEndpoint)
# ══════════════════════════════════════════════════════════════

def get_audio():
    devs = set()
    try:
        out = run_ps("Get-PnpDevice -Class AudioEndpoint | Where-Object {$_.Status -eq 'OK'} | Select-Object -ExpandProperty FriendlyName")
        for line in out.splitlines():
            nm = line.strip()
            if nm: devs.add(nm)
    except Exception: pass
    return devs

def monitor_audio():
    global st_audio; st_audio = get_audio()
    print(f"  🔊 Audio: {list(st_audio) or 'None'}")
    while True:
        time.sleep(AUDIO_POLL); cur = get_audio()
        with wb_lock:
            for dev in cur - st_audio:
                ts,d,t = now_ts()
                dtype = ("Headphones" if any(k in dev.lower() for k in ["head","ear","airpod","bud"]) else
                         "Microphone" if any(k in dev.lower() for k in ["mic","input","record"]) else "Speaker")
                write_row("Audio",[next_id(),ts,d,t,"CONNECTED",dev,dtype,"?",f"{dev} connected"],"AUDIO IN")
                log_all("AUDIO IN","Audio Connected",dev,dtype); print(f"[{ts}] 🔊 AUDIO IN : {dev}")
            for dev in st_audio - cur:
                ts,d,t = now_ts()
                write_row("Audio",[next_id(),ts,d,t,"DISCONNECTED",dev,"?","?",f"{dev} removed"],"AUDIO OUT")
                log_all("AUDIO OUT","Audio Removed",dev); print(f"[{ts}] 🔊 AUDIO OUT: {dev}")
        st_audio = cur

def get_mics():
    mics = set()
    try:
        out = run_ps("Get-PnpDevice -Class AudioEndpoint | Where-Object {$_.Status -eq 'OK' -and ($_.FriendlyName -like '*Mic*' -or $_.FriendlyName -like '*Microphone*' -or $_.FriendlyName -like '*Recording*')} | Select-Object -ExpandProperty FriendlyName")
        for line in out.splitlines():
            nm = line.strip()
            if nm: mics.add(nm)
    except Exception: pass
    return mics

def monitor_microphones():
    global st_mic; st_mic = get_mics()
    print(f"  🎙️  Mics: {list(st_mic) or 'None'}")
    while True:
        time.sleep(AUDIO_POLL); cur = get_mics()
        with wb_lock:
            for dev in cur - st_mic:
                ts,d,t = now_ts(); dtype = "USB Mic" if "usb" in dev.lower() else "Microphone"
                write_row("Microphone",[next_id(),ts,d,t,"CONNECTED",dev,dtype,f"Mic in: {dev}"],"MIC IN")
                log_all("MIC IN","Mic Connected",dev); print(f"[{ts}] 🎙️  MIC IN  : {dev}")
            for dev in st_mic - cur:
                ts,d,t = now_ts()
                write_row("Microphone",[next_id(),ts,d,t,"DISCONNECTED",dev,"Microphone",f"Mic out: {dev}"],"MIC OUT")
                log_all("MIC OUT","Mic Removed",dev); print(f"[{ts}] 🎙️  MIC OUT : {dev}")
        st_mic = cur

def get_cams():
    cams = set()
    try:
        out = run_ps("Get-PnpDevice -Class Camera,Image | Where-Object {$_.Status -eq 'OK'} | Select-Object -ExpandProperty FriendlyName")
        for line in out.splitlines():
            nm = line.strip()
            if nm: cams.add(nm)
    except Exception: pass
    return cams

def monitor_cameras():
    global st_cam; st_cam = get_cams()
    print(f"  📷 Cameras: {list(st_cam) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_cams()
        with wb_lock:
            for dev in cur - st_cam:
                ts,d,t = now_ts(); dtype = "External Webcam" if "usb" in dev.lower() else "Webcam"
                write_row("Webcam",[next_id(),ts,d,t,"CONNECTED",dev,dtype,f"Camera in: {dev}"],"CAM IN")
                log_all("CAM IN","Camera Connected",dev,dtype); print(f"[{ts}] 📷 CAM IN  : {dev}")
            for dev in st_cam - cur:
                ts,d,t = now_ts()
                write_row("Webcam",[next_id(),ts,d,t,"DISCONNECTED",dev,"Camera",f"Camera out: {dev}"],"CAM OUT")
                log_all("CAM OUT","Camera Removed",dev); print(f"[{ts}] 📷 CAM OUT : {dev}")
        st_cam = cur

def get_printers():
    printers = {}
    try:
        out = run_ps("Get-Printer | Select-Object Name,PortName,DriverName,PrinterStatus | Format-List")
        info = {}
        for line in out.splitlines():
            line = line.strip()
            if   line.startswith("Name"):          info["name"]   = line.split(":",1)[-1].strip()
            elif line.startswith("PortName"):      info["port"]   = line.split(":",1)[-1].strip()
            elif line.startswith("DriverName"):    info["driver"] = line.split(":",1)[-1].strip()
            elif line.startswith("PrinterStatus"):
                info["status"] = line.split(":",1)[-1].strip()
                if info.get("name"):
                    printers[info["name"]] = {"port":info.get("port","?"),"driver":info.get("driver","?"),"status":info.get("status","?")}
                info = {}
    except Exception: pass
    return printers

def monitor_printers():
    global st_printers; st_printers = get_printers()
    print(f"  🖨️  Printers: {list(st_printers) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_printers()
        with wb_lock:
            for nm in set(cur) - set(st_printers):
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Printers",[next_id(),ts,d,t,"ADDED",nm,info["port"],info["driver"],info["status"]],"PRINTER ADD")
                log_all("PRINTER ADD","Printer Added",nm,info["port"]); print(f"[{ts}] 🖨️  PRTR ADD: {nm}")
            for nm in set(st_printers) - set(cur):
                old = st_printers.get(nm,{}); ts,d,t = now_ts()
                write_row("Printers",[next_id(),ts,d,t,"REMOVED",nm,old.get("port","?"),old.get("driver","?"),"Removed"],"PRINTER REMOVE")
                log_all("PRINTER REMOVE","Printer Removed",nm); print(f"[{ts}] 🖨️  PRTR REM: {nm}")
        st_printers = cur

def get_monitors():
    mons = {}
    try:
        out = run_ps("Get-PnpDevice -Class Monitor | Where-Object {$_.Status -eq 'OK'} | Select-Object FriendlyName,DeviceID | Format-List")
        name = None
        for line in out.splitlines():
            line = line.strip()
            if line.startswith("FriendlyName"): name = line.split(":",1)[-1].strip()
            elif line.startswith("DeviceID") and name:
                did = line.split(":",1)[-1].strip()
                port = "HDMI" if "HDMI" in did else "DisplayPort" if "DP" in did else "VGA" if "VGA" in did else "?"
                mons[name] = {"res":"?","port":port,"serial":did[-12:]}; name = None
        res_out = run_ps("Get-CimInstance Win32_VideoController | Select-Object CurrentHorizontalResolution,CurrentVerticalResolution | Format-List")
        w = h = ""
        for line in res_out.splitlines():
            if "Horizontal" in line: w = line.split(":")[-1].strip()
            elif "Vertical"  in line: h = line.split(":")[-1].strip()
        res = f"{w}x{h}" if w and h else "?"
        for nm in mons: mons[nm]["res"] = res
    except Exception: pass
    return mons

def monitor_monitors():
    global st_monitors; st_monitors = get_monitors()
    print(f"  📺 Monitors: {list(st_monitors) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_monitors()
        with wb_lock:
            for nm in set(cur) - set(st_monitors):
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Monitors",[next_id(),ts,d,t,"CONNECTED",nm,info["res"],info["port"],info["serial"]],"MONITOR IN")
                log_all("MONITOR IN","Monitor Connected",nm,f"{info['res']} via {info['port']}")
                print(f"[{ts}] 📺 MON IN  : {nm} ({info['res']})")
            for nm in set(st_monitors) - set(cur):
                old = st_monitors.get(nm,{}); ts,d,t = now_ts()
                write_row("Monitors",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("res","?"),old.get("port","?"),old.get("serial","?")],"MONITOR OUT")
                log_all("MONITOR OUT","Monitor Disconnected",nm); print(f"[{ts}] 📺 MON OUT : {nm}")
        st_monitors = cur

# ══════════════════════════════════════════════════════════════
#  VPN  (Windows: Get-VpnConnection + VPN adapters)
# ══════════════════════════════════════════════════════════════

def get_vpn():
    vpns = {}
    try:
        out = run_ps("Get-VpnConnection | Where-Object {$_.ConnectionStatus -eq 'Connected'} | Select-Object Name,ServerAddress,TunnelType | Format-List")
        info = {}
        for line in out.splitlines():
            line = line.strip()
            if   line.startswith("Name"):          info["name"]   = line.split(":",1)[-1].strip()
            elif line.startswith("ServerAddress"): info["server"] = line.split(":",1)[-1].strip()
            elif line.startswith("TunnelType"):
                info["proto"] = line.split(":",1)[-1].strip()
                if info.get("name"):
                    vpns[info["name"]] = {"server":info.get("server","?"),"proto":info.get("proto","?")}
                info = {}
        for iface in psutil.net_if_stats():
            n = iface.lower()
            if any(v in n for v in ["vpn","tun","tap","ppp","nordlynx","wireguard","ipsec"]):
                if psutil.net_if_stats()[iface].isup:
                    vpns[iface] = {"server":"?","proto":"VPN Adapter"}
    except Exception: pass
    return vpns

def monitor_vpn():
    global st_vpn
    prev = get_vpn(); st_vpn = set(prev.keys()); vpn_since = {}
    print(f"  🌍 VPN: {list(st_vpn) or 'None'}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_vpn()
        with wb_lock:
            for nm in set(cur) - st_vpn:
                info = cur[nm]; ts,d,t = now_ts(); vpn_since[nm] = datetime.datetime.now()
                write_row("VPN",[next_id(),ts,d,t,"CONNECTED",nm,info["server"],info["proto"],""],"VPN CONNECT")
                log_all("VPN CONNECT","VPN Connected",nm,f"Server:{info['server']}"); print(f"[{ts}] 🌍 VPN CONN: {nm}")
            for nm in st_vpn - set(cur):
                old = prev.get(nm,{}); ts,d,t = now_ts()
                dur = str(datetime.datetime.now()-vpn_since[nm]).split(".")[0] if nm in vpn_since else "?"
                write_row("VPN",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("server","?"),old.get("proto","?"),dur],"VPN DISCONNECT")
                log_all("VPN DISCONNECT","VPN Disconnected",nm,"",dur); print(f"[{ts}] 🌍 VPN DISC: {nm}")
        st_vpn = set(cur); prev = cur

def get_hotspot():
    hs = set()
    try:
        for q in [
            "Get-NetAdapter | Where-Object {$_.Status -eq 'Up' -and ($_.InterfaceDescription -like '*Hosted*' -or $_.InterfaceDescription -like '*Hotspot*' -or $_.InterfaceDescription -like '*Bluetooth PAN*')} | Select-Object -ExpandProperty Name",
            "Get-NetAdapter | Where-Object {$_.InterfaceDescription -like '*RNDIS*' -and $_.Status -eq 'Up'} | Select-Object -ExpandProperty Name",
        ]:
            out = run_ps(q)
            for line in out.splitlines():
                nm = line.strip()
                if nm: hs.add(nm)
    except Exception: pass
    return hs

def monitor_hotspot():
    global st_hotspot; prev = get_hotspot(); st_hotspot = prev
    print(f"  📱 Hotspot: {list(prev) or 'None'}")
    while True:
        time.sleep(NETWORK_POLL); cur = get_hotspot()
        with wb_lock:
            for dev in cur - prev:
                ts,d,t = now_ts(); dtype = "USB Tethering" if "RNDIS" in dev else "Mobile Hotspot"
                write_row("Mobile Hotspot",[next_id(),ts,d,t,"CONNECTED",dev,dtype,"?","Mobile connected"],"HOTSPOT ON")
                log_all("HOTSPOT ON","Hotspot Connected",dev,dtype); print(f"[{ts}] 📱 HOTSPOT+: {dev}")
            for dev in prev - cur:
                ts,d,t = now_ts()
                write_row("Mobile Hotspot",[next_id(),ts,d,t,"DISCONNECTED",dev,"?","?","Mobile disconnected"],"HOTSPOT OFF")
                log_all("HOTSPOT OFF","Hotspot Removed",dev); print(f"[{ts}] 📱 HOTSPOT-: {dev}")
        prev = cur; st_hotspot = cur

def get_net_drives():
    drives = {}
    try:
        out = run_ps("Get-PSDrive -PSProvider FileSystem | Where-Object {$_.DisplayRoot -like '\\\\*'} | Select-Object Name,DisplayRoot | Format-List")
        info = {}
        for line in out.splitlines():
            line = line.strip()
            if line.startswith("Name"):        info["name"] = line.split(":",1)[-1].strip()
            elif line.startswith("DisplayRoot"):
                info["unc"] = line.split(":",1)[-1].strip()
                if info.get("name"):
                    srv = info["unc"].split("\\")[2] if "\\" in info["unc"] else "?"
                    drives[info["name"]] = {"unc":info["unc"],"server":srv}
                info = {}
    except Exception: pass
    return drives

def monitor_net_drives():
    global st_netdrives; prev = get_net_drives(); st_netdrives = set(prev.keys())
    print(f"  🗂️  Net Drives: {list(st_netdrives) or 'None'}")
    while True:
        time.sleep(DRIVE_POLL); cur = get_net_drives()
        with wb_lock:
            for nm in set(cur) - st_netdrives:
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Network Drives",[next_id(),ts,d,t,"MAPPED",nm,info["unc"],info["server"],"Connected"],"NET DRIVE MAP")
                log_all("NET DRIVE MAP","Net Drive Mapped",nm,info["unc"]); print(f"[{ts}] 🗂️  NETDRV+: {nm}")
            for nm in st_netdrives - set(cur):
                old = prev.get(nm,{}); ts,d,t = now_ts()
                write_row("Network Drives",[next_id(),ts,d,t,"UNMAPPED",nm,old.get("unc","?"),old.get("server","?"),"Disconnected"],"NET DRIVE UNMAP")
                log_all("NET DRIVE UNMAP","Net Drive Unmapped",nm); print(f"[{ts}] 🗂️  NETDRV-: {nm}")
        st_netdrives = set(cur); prev = cur

def get_controllers():
    ctrls = {}
    try:
        out = run_ps(
            "Get-PnpDevice -Class HIDClass | Where-Object {$_.Status -eq 'OK' -and "
            "($_.FriendlyName -like '*Controller*' -or $_.FriendlyName -like '*Gamepad*' -or "
            "$_.FriendlyName -like '*Xbox*' -or $_.FriendlyName -like '*PlayStation*' -or "
            "$_.FriendlyName -like '*DualShock*' -or $_.FriendlyName -like '*DualSense*')} "
            "| Select-Object FriendlyName,DeviceID | Format-List"
        )
        name = None
        for line in out.splitlines():
            line = line.strip()
            if line.startswith("FriendlyName"): name = line.split(":",1)[-1].strip()
            elif line.startswith("DeviceID") and name:
                did = line.split(":",1)[-1].strip()
                ctrls[name] = {"type":"Controller","id":did[-16:]}; name = None
    except Exception: pass
    return ctrls

def monitor_controllers():
    global st_ctrl; prev = get_controllers(); st_ctrl = set(prev.keys())
    print(f"  🎮 Controllers: {list(st_ctrl) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_controllers()
        with wb_lock:
            for nm in set(cur) - st_ctrl:
                info = cur[nm]; ts,d,t = now_ts()
                write_row("Controllers",[next_id(),ts,d,t,"CONNECTED",nm,info["type"],info["id"],"Ready"],"CTRL CONNECT")
                log_all("CTRL CONNECT","Controller Connected",nm,info["type"]); print(f"[{ts}] 🎮 CTRL IN : {nm}")
            for nm in st_ctrl - set(cur):
                old = prev.get(nm,{}); ts,d,t = now_ts()
                write_row("Controllers",[next_id(),ts,d,t,"DISCONNECTED",nm,old.get("type","?"),old.get("id","?"),"Disconnected"],"CTRL DISCONNECT")
                log_all("CTRL DISCONNECT","Controller Removed",nm); print(f"[{ts}] 🎮 CTRL OUT: {nm}")
        st_ctrl = set(cur); prev = cur

def get_remote():
    sessions = {}
    try:
        out = subprocess.check_output(["query","session"],text=True,stderr=subprocess.DEVNULL,timeout=8)
        for line in out.splitlines()[1:]:
            if any(k in line.lower() for k in ["rdp","remote","teamviewer","anydesk"]):
                parts = line.split()
                if len(parts) >= 3:
                    sessions[parts[1] if len(parts)>1 else "?"] = {"session":parts[0],"type":"RDP"}
    except Exception: pass
    return sessions

def monitor_remote():
    global st_rdp; prev = get_remote(); st_rdp = set(prev.keys()); rdp_since = {}
    while True:
        time.sleep(DEVICE_POLL); cur = get_remote()
        with wb_lock:
            for nm in set(cur) - st_rdp:
                ts,d,t = now_ts(); rdp_since[nm] = datetime.datetime.now()
                write_row("Remote Sessions",[next_id(),ts,d,t,"CONNECTED",nm,"?","RDP",""],"RDP IN")
                log_all("RDP IN","RDP Session Started",nm); print(f"[{ts}] 🔐 RDP IN  : {nm}")
            for nm in st_rdp - set(cur):
                ts,d,t = now_ts(); dur = str(datetime.datetime.now()-rdp_since[nm]).split(".")[0] if nm in rdp_since else "?"
                write_row("Remote Sessions",[next_id(),ts,d,t,"DISCONNECTED",nm,"?","RDP",dur],"RDP OUT")
                log_all("RDP OUT","RDP Session Ended",nm,"",dur); print(f"[{ts}] 🔐 RDP OUT : {nm}")
        st_rdp = set(cur); prev = cur

def monitor_security():
    prev_fw = prev_def = prev_uac = None
    print("  🛡️  Security: Monitoring")
    while True:
        time.sleep(60)
        for check, cmd, label, key in [
            ("fw",  "(Get-NetFirewallProfile -Profile Domain,Public,Private | Select-Object -ExpandProperty Enabled) -join ','", "FIREWALL CHANGED", "Windows Firewall"),
            ("def", "(Get-MpComputerStatus).AntivirusEnabled", "DEFENDER CHANGED", "Windows Defender"),
            ("uac", "(Get-ItemProperty HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System).EnableLUA", "UAC CHANGED", "User Account Control"),
        ]:
            try:
                out = run_ps(cmd).strip()
                prev = {"fw":prev_fw,"def":prev_def,"uac":prev_uac}[check]
                if out != prev and prev is not None:
                    ts,d,t = now_ts()
                    cat = "SECURITY OK" if (check=="def" and out=="True") else "SECURITY"
                    with wb_lock:
                        write_row("Security",[next_id(),ts,d,t,label,key,out,f"Old:{prev}→New:{out}"],cat)
                        log_all(cat,f"{key} Changed",key,out)
                        print(f"[{ts}] 🛡️  {label}: {out}")
                if   check=="fw":  prev_fw  = out
                elif check=="def": prev_def = out
                elif check=="uac": prev_uac = out
            except Exception: pass

def monitor_users():
    prev = {u.name: u for u in psutil.users()}
    print(f"  👤 Users: {list(prev)}")
    while True:
        time.sleep(10)
        try: cur = {u.name: u for u in psutil.users()}
        except Exception: continue
        with wb_lock:
            for nm in set(cur) - set(prev):
                u = cur[nm]; ts,d,t = now_ts()
                started = datetime.datetime.fromtimestamp(u.started).strftime("%Y-%m-%d %H:%M:%S")
                write_row("Users",[next_id(),ts,d,t,"LOGIN",nm,u.terminal or "?",f"Session:{started}",pc_name],"USER LOGIN")
                log_all("USER LOGIN","Login",nm,u.terminal or "?"); print(f"[{ts}] 👤 LOGIN   : {nm}")
            for nm in set(prev) - set(cur):
                ts,d,t = now_ts()
                write_row("Users",[next_id(),ts,d,t,"LOGOUT",nm,prev[nm].terminal or "?","",pc_name],"USER LOGOUT")
                log_all("USER LOGOUT","Logout",nm); print(f"[{ts}] 👤 LOGOUT  : {nm}")
        prev = cur

def get_input_devs(cls):
    devs = set()
    try:
        out = run_ps(f"Get-PnpDevice -Class {cls} | Where-Object {{$_.Status -eq 'OK' -and $_.InstanceId -like 'USB*'}} | Select-Object -ExpandProperty FriendlyName")
        for line in out.splitlines():
            nm = line.strip()
            if nm: devs.add(nm)
    except Exception: pass
    return devs

def monitor_keyboards():
    global st_kb; st_kb = get_input_devs("Keyboard")
    print(f"  ⌨️  Keyboards: {list(st_kb) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_input_devs("Keyboard")
        with wb_lock:
            for dev in cur - st_kb:
                ts,d,t = now_ts(); iface = "USB" if "USB" in dev else "Wireless"
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"CONNECTED",dev,"Keyboard",iface,"?"],"KB CONNECT")
                log_all("KB CONNECT","Keyboard Connected",dev,iface); print(f"[{ts}] ⌨️  KB IN   : {dev}")
            for dev in st_kb - cur:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"DISCONNECTED",dev,"Keyboard","?","?"],"KB DISCONNECT")
                log_all("KB DISCONNECT","Keyboard Disconnected",dev); print(f"[{ts}] ⌨️  KB OUT  : {dev}")
        st_kb = cur

def monitor_mice():
    global st_mouse; st_mouse = get_input_devs("Mouse")
    print(f"  🖱️  Mice: {list(st_mouse) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_input_devs("Mouse")
        with wb_lock:
            for dev in cur - st_mouse:
                ts,d,t = now_ts(); iface = "USB" if "USB" in dev else "Wireless"
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"CONNECTED",dev,"Mouse",iface,"?"],"MOUSE CONNECT")
                log_all("MOUSE CONNECT","Mouse Connected",dev,iface); print(f"[{ts}] 🖱️  MOUSE IN : {dev}")
            for dev in st_mouse - cur:
                ts,d,t = now_ts()
                write_row("Keyboard Mouse",[next_id(),ts,d,t,"DISCONNECTED",dev,"Mouse","?","?"],"MOUSE DISCONNECT")
                log_all("MOUSE DISCONNECT","Mouse Disconnected",dev); print(f"[{ts}] 🖱️  MOUSE OUT: {dev}")
        st_mouse = cur

def get_power_plan():
    try:
        out = run_ps("(powercfg /getactivescheme)").strip()
        m  = re.search(r"\((.+?)\)", out); g = re.search(r"([0-9a-f\-]{36})", out, re.IGNORECASE)
        return (m.group(1).strip() if m else out), (g.group(1) if g else "?")
    except Exception: return "?","?"

def monitor_power_plan():
    global st_plan; st_plan,_ = get_power_plan()
    print(f"  ⚙️  Power Plan: {st_plan}")
    while True:
        time.sleep(SERVICE_POLL); plan,guid = get_power_plan()
        if plan and plan != st_plan:
            ts,d,t = now_ts()
            with wb_lock:
                write_row("Power Plan",[next_id(),ts,d,t,"CHANGED",plan,guid,st_plan,f"Switched to: {plan}"],"PLAN CHANGED")
                log_all("PLAN CHANGED","Power Plan Changed",plan,f"Was:{st_plan}"); print(f"[{ts}] ⚙️  PLAN: {st_plan}→{plan}")
            st_plan = plan

def get_svc_states():
    states = {}
    try:
        names = ",".join(f'"{s}"' for s in WATCHED_SERVICES)
        out   = run_ps(f"Get-Service -Name {names} -ErrorAction SilentlyContinue | Select-Object Name,DisplayName,Status | Format-List", timeout=15)
        info  = {}
        for line in out.splitlines():
            line = line.strip()
            if   line.startswith("Name"):        info["name"]   = line.split(":",1)[-1].strip()
            elif line.startswith("DisplayName"): info["disp"]   = line.split(":",1)[-1].strip()
            elif line.startswith("Status"):
                info["status"] = line.split(":",1)[-1].strip()
                if info.get("name"):
                    states[info["name"]] = {"disp":info.get("disp","?"),"status":info["status"]}
                info = {}
    except Exception: pass
    return states

def monitor_services():
    global st_svc; st_svc = get_svc_states()
    print(f"  🔧 Services: {len(st_svc)} monitored")
    while True:
        time.sleep(SERVICE_POLL); cur = get_svc_states()
        with wb_lock:
            for nm, info in cur.items():
                old = st_svc.get(nm,{})
                if old.get("status") and old["status"] != info["status"]:
                    ts,d,t = now_ts(); cat = "SVC START" if info["status"]=="Running" else "SVC STOP"
                    write_row("Services",[next_id(),ts,d,t,info["status"],nm,info["disp"],old["status"],info["status"]],cat)
                    log_all(cat,f"Service {info['status']}",nm,f"{old['status']}→{info['status']}")
                    print(f"[{ts}] 🔧 SVC {info['status']:12}: {nm}")
        st_svc = cur

def monitor_perf():
    if not LOG_PERF: return
    print(f"  📊 Perf: CPU>{CPU_SPIKE_THRESH}% / RAM>{RAM_SPIKE_THRESH}%")
    was_cpu = was_ram = False
    while True:
        time.sleep(PERF_POLL); cpu = psutil.cpu_percent(interval=2); ram = psutil.virtual_memory().percent
        ts,d,t = now_ts()
        with wb_lock:
            if cpu >= CPU_SPIKE_THRESH and not was_cpu:
                try:
                    top = max(psutil.process_iter(["name","cpu_percent"]),key=lambda p:p.info["cpu_percent"] or 0)
                    top_name = f"{top.info['name']} ({top.info['cpu_percent']:.0f}%)"
                except Exception: top_name = "?"
                write_row("System Perf",[next_id(),ts,d,t,"CPU SPIKE",f"{cpu:.1f}%",f"{ram:.1f}%",top_name,f"CPU spike:{cpu:.1f}%"],"CPU SPIKE")
                log_all("CPU SPIKE","CPU Spike",f"{cpu:.1f}%",top_name); print(f"[{ts}] 🔴 CPU SPIKE: {cpu:.1f}%")
            if ram >= RAM_SPIKE_THRESH and not was_ram:
                write_row("System Perf",[next_id(),ts,d,t,"RAM SPIKE",f"{cpu:.1f}%",f"{ram:.1f}%","?",f"RAM spike:{ram:.1f}%"],"RAM SPIKE")
                log_all("RAM SPIKE","RAM Spike",f"{ram:.1f}%",f"CPU:{cpu:.1f}%"); print(f"[{ts}] 🔴 RAM SPIKE: {ram:.1f}%")
        was_cpu = (cpu >= CPU_SPIKE_THRESH); was_ram = (ram >= RAM_SPIKE_THRESH)

def get_smartcards():
    cards = set()
    try:
        out = run_ps("Get-PnpDevice -Class SmartCardReader | Where-Object {$_.Status -eq 'OK'} | Select-Object -ExpandProperty FriendlyName")
        for line in out.splitlines():
            nm = line.strip()
            if nm: cards.add(nm)
    except Exception: pass
    return cards

def monitor_smartcard():
    global st_card; st_card = get_smartcards()
    print(f"  🔑 SmartCard: {list(st_card) or 'None'}")
    while True:
        time.sleep(DEVICE_POLL); cur = get_smartcards()
        with wb_lock:
            for dev in cur - st_card:
                ts,d,t = now_ts()
                write_row("Smart Card",[next_id(),ts,d,t,"CONNECTED",dev,"Smart Card Reader",f"Reader in: {dev}"],"CARD IN")
                log_all("CARD IN","Smart Card Connected",dev); print(f"[{ts}] 🔑 CARD IN : {dev}")
            for dev in st_card - cur:
                ts,d,t = now_ts()
                write_row("Smart Card",[next_id(),ts,d,t,"DISCONNECTED",dev,"Smart Card Reader",f"Reader out: {dev}"],"CARD OUT")
                log_all("CARD OUT","Smart Card Removed",dev); print(f"[{ts}] 🔑 CARD OUT: {dev}")
        st_card = cur

def get_display_settings():
    settings = {}
    try:
        out = run_ps("(Get-ItemProperty 'HKCU:\\Software\\Microsoft\\Windows\\CurrentVersion\\CloudStore\\Store\\DefaultAccount\\Current\\default$windows.data.bluelightreduction.bluelightreductionstate\\windows.data.bluelightreduction.bluelightreductionstate' -ErrorAction SilentlyContinue).Data").strip()
        settings["NightLight"] = "ON" if out and len(out) > 10 else "OFF"
    except Exception: pass
    try:
        out = run_ps("(Get-ItemProperty 'HKCU:\\Software\\Microsoft\\Windows\\CurrentVersion\\VideoSettings' -ErrorAction SilentlyContinue).EnableHDRForVideos").strip()
        settings["HDR"] = f"EnableHDR:{out}"
    except Exception: pass
    return settings

def monitor_display_mode():
    prev = get_display_settings(); print(f"  🌙 Display Mode: {prev}")
    while True:
        time.sleep(SERVICE_POLL); cur = get_display_settings()
        with wb_lock:
            for key, val in cur.items():
                old_val = prev.get(key)
                if old_val and val != old_val:
                    ts,d,t = now_ts()
                    write_row("Display Mode",[next_id(),ts,d,t,"CHANGED",key,old_val,val,f"{key}: {old_val}→{val}"],"DISPLAY CHANGE")
                    log_all("DISPLAY CHANGE","Display Mode Changed",key,f"{old_val}→{val}")
                    print(f"[{ts}] 🌙 DISPLAY : {key} {old_val}→{val}")
        prev = cur

def monitor_clipboard():
    global clip_last
    if not LOG_CLIPBOARD: return
    print("  📋 Clipboard: ENABLED")
    while True:
        time.sleep(3)
        try:
            import win32clipboard
            win32clipboard.OpenClipboard()
            try:    text = win32clipboard.GetClipboardData()
            except: text = ""
            finally: win32clipboard.CloseClipboard()
            if text and text != clip_last:
                snippet = text[:100].replace("\n"," ").replace("\r",""); ts,d,t = now_ts()
                with wb_lock:
                    write_row("Clipboard",[next_id(),ts,d,t,snippet,len(text),"Content changed"],"CLIPBOARD")
                clip_last = text
        except Exception: pass

def save_wb():
    try:
        os.makedirs(os.path.dirname(LOG_FILE) if os.path.dirname(LOG_FILE) else ".", exist_ok=True)
        wb.save(LOG_FILE)
    except Exception as e: print(f"  ⚠️  Save error: {e}")

def auto_save():
    while True:
        time.sleep(SAVE_EVERY)
        with wb_lock: save_wb()
        print(f"  💾 Auto-saved @ {datetime.datetime.now().strftime('%H:%M:%S')}")

# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

def main():
    global wb
    print("╔"+"═"*62+"╗")
    print("║   PC ACTIVITY LOGGER — WINDOWS EDITION                   ║")
    print("║   30 categories | Optimised for Windows 10 / 11          ║")
    print("╚"+"═"*62+"╝")
    wb = init_excel()
    with wb_lock: log_startup(); save_wb()

    monitors = [
        ("Apps",        monitor_apps),         ("USB",         monitor_usb),
        ("WiFi",        monitor_wifi),          ("Bluetooth",   monitor_bluetooth),
        ("Network",     monitor_network),       ("Internet",    monitor_internet),
        ("Battery",     monitor_battery),       ("Screen",      monitor_screen),
        ("Drives",      monitor_drives),        ("Audio",       monitor_audio),
        ("Microphone",  monitor_microphones),   ("Camera",      monitor_cameras),
        ("Printers",    monitor_printers),      ("Monitors",    monitor_monitors),
        ("VPN",         monitor_vpn),           ("Hotspot",     monitor_hotspot),
        ("NetDrives",   monitor_net_drives),    ("Controllers", monitor_controllers),
        ("Remote",      monitor_remote),        ("Security",    monitor_security),
        ("Users",       monitor_users),         ("Keyboards",   monitor_keyboards),
        ("Mice",        monitor_mice),          ("PowerPlan",   monitor_power_plan),
        ("Services",    monitor_services),      ("PerfMonitor", monitor_perf),
        ("SmartCard",   monitor_smartcard),     ("DisplayMode", monitor_display_mode),
        ("Clipboard",   monitor_clipboard),     ("AutoSave",    auto_save),
    ]
    for name, fn in monitors:
        threading.Thread(target=fn, daemon=True, name=name).start()

    print(f"\n  📁 Log : {LOG_FILE}")
    print(f"  💾 Auto-save every {SAVE_EVERY}s | {len(monitors)-1} monitors active")
    print("  ⏹️  Press Ctrl+C to stop\n")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        print("\n  Stopping..."); 
        with wb_lock: log_shutdown()
        print("  ✅ Done.")

if __name__ == "__main__":
    if platform.system() != "Windows":
        print("⚠️  This file is for Windows only. Use pc_logger_mac.py or pc_logger_linux.py instead.")
        sys.exit(1)
    main()
